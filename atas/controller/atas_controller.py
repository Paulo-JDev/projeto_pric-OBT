# atas/controller/atas_controller.py
from PyQt6.QtWidgets import QFileDialog, QMessageBox, QHeaderView, QMenu
from PyQt6.QtGui import QStandardItem, QBrush, QColor, QFont
from PyQt6.QtCore import Qt
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os
import shutil
import json

from atas.model.atas_model import AtasModel
from atas.model.atas_model import Base, engine
from utils.icon_loader import icon_manager 
from atas.view.ata_details_dialog import AtaDetailsDialog
from atas.controller.controller_fiscal_ata import save_fiscalizacao_ata

from integration.model.trello_model import TrelloModel
from integration.controller.trello_individual_controller import TrelloIndividualController, TrelloSyncWorker

class AtasController:
    def __init__(self, model: AtasModel, view):
        self.model = model
        self.view = view
        self._connect_signals()
        #self.load_initial_data()
        self.check_db_status_and_load_data()

    def _connect_signals(self):
        self.view.delete_button.clicked.connect(self.delete_selected_ata)
        self.view.add_button.clicked.connect(self.show_add_ata_dialog)

        # --- SIGNALS DO MENU "PLANILHA" (antigo "DADOS") ---
        self.view.import_action.triggered.connect(self.import_data)
        self.view.export_completo_action.triggered.connect(self.generate_excel_report)

        self.view.export_bi_action.triggered.connect(self.generate_bi_export)

        self.view.template_vazio_action.triggered.connect(self.generate_empty_template)
        self.view.export_para_importacao_action.triggered.connect(self.export_for_reimport)

        # --- SIGNALS DO NOVO MENU "DB" ---
        self.view.change_db_location_action.triggered.connect(self.change_database_location)
        self.view.export_main_json_action.triggered.connect(self.export_main_data_json)
        self.view.import_main_json_action.triggered.connect(self.import_main_data_json)
        self.view.export_complementary_json_action.triggered.connect(self.export_complementary_data_json)
        self.view.import_complementary_json_action.triggered.connect(self.import_complementary_data_json)

        self.view.table_view.doubleClicked.connect(self.show_details_on_double_click)
        self.view.table_view.customContextMenuRequested.connect(self.show_context_menu)
        self.view.preview_table.doubleClicked.connect(self.show_details_on_preview_double_click)
        self.view.refresh_preview_button.clicked.connect(self.populate_previsualization_table)
        self.view.refresh_table_button.clicked.connect(self.load_initial_data)

    def check_db_status_and_load_data(self):
        """Verifica o status do DB e carrega os dados ou inicia o processo de migração."""

        # 🚦 CASO 1: DB desatualizado, mas ainda acessível para exportação
        if not self.model.db_initialized and getattr(self.model, "allow_raw_export", False):
            # 🔒 Bloqueia tudo que modifica dados
            self.view.add_button.setEnabled(False)
            self.view.delete_button.setEnabled(False)
            self.view.planilha_button.setEnabled(False)
            self.view.refresh_table_button.setEnabled(False)
            self.view.refresh_preview_button.setEnabled(False)
            self.view.preview_table.setEnabled(False)
            self.view.table_view.setEnabled(False)

            # 🟢 Habilita só o necessário pra migração
            self.view.export_main_json_action.setEnabled(True)
            self.view.export_complementary_json_action.setEnabled(True)
            self.view.import_main_json_action.setEnabled(False)
            self.view.import_complementary_json_action.setEnabled(False)
            self.view.change_db_location_action.setEnabled(True)

            # 🧭 Explica o que deve ser feito
            QMessageBox.warning(
                self.view,
                "Banco de Dados Antigo Detectado",
                (
                    "O banco de dados atual foi feito em uma versão antiga do programa.\n\n"
                    "Você pode exportar agora todos os seus dados.\n\n"
                    "➡️ **PASSOS PARA MIGRAR:**\n"
                    "1. Use o menu 'DB' para **Exportar Dados Principais (JSON)** e **Exportar Dados Complementares (JSON)**.\n"
                    "2. Feche o programa.\n"
                    "3. Exclua ou renomeie o arquivo `atas_controle.db` antigo.\n"
                    "4. Reinicie o programa — ele criará um novo DB atualizado.\n"
                    "5. Depois, use o menu 'DB' para **Importar Seus Dados JSON**.\n\n"
                    "⚠️ Nenhum dado será apagado enquanto você não criar o novo banco."
                )
            )
            print("⚠️ Banco antigo — exportação via sqlite3 habilitada.")
            return

        # 🚦 CASO 2: DB completamente inválido (nem exportável)
        elif not self.model.db_initialized and not getattr(self.model, "allow_raw_export", False):
            # trava tudo
            self.view.add_button.setEnabled(False)
            self.view.delete_button.setEnabled(False)
            self.view.planilha_button.setEnabled(False)
            self.view.refresh_table_button.setEnabled(False)
            self.view.refresh_preview_button.setEnabled(False)
            self.view.preview_table.setEnabled(False)
            self.view.table_view.setEnabled(False)

            # permite mudar local do DB, mas sem exportar
            self.view.export_main_json_action.setEnabled(False)
            self.view.export_complementary_json_action.setEnabled(False)
            self.view.import_main_json_action.setEnabled(False)
            self.view.import_complementary_json_action.setEnabled(False)
            self.view.change_db_location_action.setEnabled(True)

            QMessageBox.critical(
                self.view,
                "Erro Crítico no Banco de Dados",
                (
                    "O banco de dados não pôde ser aberto ou está corrompido.\n\n"
                    "Por favor, selecione um novo local de banco ou crie um novo DB vazio via menu 'DB > Mudar Local do DB'."
                )
            )
            print("❌ DB corrompido — exportação indisponível.")
            return

        # 🚦 CASO 3: Tudo certo (schema atualizado)
        else:
            # habilita tudo normalmente
            self.view.add_button.setEnabled(True)
            self.view.delete_button.setEnabled(True)
            self.view.planilha_button.setEnabled(True)
            self.view.refresh_table_button.setEnabled(True)
            self.view.refresh_preview_button.setEnabled(True)
            self.view.preview_table.setEnabled(True)
            self.view.table_view.setEnabled(True)

            # habilita todas as opções JSON e DB
            self.view.export_main_json_action.setEnabled(True)
            self.view.export_complementary_json_action.setEnabled(True)
            self.view.import_main_json_action.setEnabled(True)
            self.view.import_complementary_json_action.setEnabled(True)
            self.view.change_db_location_action.setEnabled(True)

            # carrega dados
            self.load_initial_data()
            print("✅ Banco OK — interface reativada e dados carregados.")

     # --- FUNÇÕES PARA GERENCIAR O LOCAL DO DB ---
    def change_database_location(self):
        """Permite ao usuário escolher um novo local para o banco de dados."""
        from pathlib import Path

        new_folder = QFileDialog.getExistingDirectory(
            self.view, 
            "Selecione a Nova Pasta para o Banco de Dados",
            str(Path.home())
        )

        if not new_folder:
            return

        new_db_file_path = Path(new_folder) / "atas_controle.db"
        current_db_file_path = self.model.get_current_db_path()

        if new_db_file_path.exists():
            reply = QMessageBox.question(
                self.view,
                "Banco de Dados Existente",
                f"Já existe um banco de dados em:\n{new_db_file_path}\n\n"
                "Deseja usar este banco existente?\n\n"
                "• SIM: Usar o banco existente\n"
                "• NÃO: Copiar o banco atual para lá (substituindo)",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
            )

            if reply == QMessageBox.StandardButton.Cancel:
                return
            elif reply == QMessageBox.StandardButton.No:
                try:
                    shutil.copy2(current_db_file_path, new_db_file_path)
                    QMessageBox.information(
                        self.view, 
                        "Banco Copiado", 
                        f"Banco de dados copiado para:\n{new_db_file_path}"
                    )
                except Exception as e:
                    QMessageBox.critical(
                        self.view, 
                        "Erro ao Copiar", 
                        f"Não foi possível copiar o banco:\n{e}"
                    )
                    return
        else:
            reply = QMessageBox.question(
                self.view,
                "Configurar Novo Local",
                f"O que deseja fazer?\n\n"
                f"• SIM: Copiar o banco atual para {new_folder}\n"
                f"• NÃO: Criar um banco vazio no novo local",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel
            )

            if reply == QMessageBox.StandardButton.Cancel:
                return
            elif reply == QMessageBox.StandardButton.Yes:
                try:
                    shutil.copy2(current_db_file_path, new_db_file_path)
                except Exception as e:
                    QMessageBox.critical(
                        self.view, 
                        "Erro ao Copiar", 
                        f"Não foi possível copiar o banco:\n{e}"
                    )
                    return

        success = self.model.change_database_path(str(new_folder))

        if success:
            QMessageBox.information(
                self.view,
                "Sucesso",
                f"Banco de dados alterado para:\n{new_db_file_path}\n\n"
                "A configuração foi salva e será mantida nas próximas execuções."
            )
            self.check_db_status_and_load_data() # Re-verifica o status do DB e recarrega
        else:
            QMessageBox.critical(
                self.view,
                "Erro",
                "Não foi possível alterar o local do banco de dados."
            )

    # --- NOVAS FUNÇÕES DE EXPORTAÇÃO/IMPORTAÇÃO JSON ---

    def export_main_data_json(self):
        """Exporta os dados principais (tabela Ata) para um arquivo JSON."""
        file_path, _ = QFileDialog.getSaveFileName(
            self.view, "Exportar Dados Principais (JSON)", "atas_principais.json",
            "JSON Files (*.json);;All Files (*)"
        )
        if not file_path: return

        success, data = self.model.export_main_data_to_json()
        if success:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=4, ensure_ascii=False)
                QMessageBox.information(self.view, "Sucesso", f"Dados principais exportados para:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self.view, "Erro", f"Não foi possível salvar o arquivo JSON:\n{e}")
        else:
            QMessageBox.critical(self.view, "Erro na Exportação", f"Ocorreu um erro ao exportar os dados principais:\n{data}")

    def import_main_data_json(self):
        """Importa os dados principais (tabela Ata) de um arquivo JSON."""
        file_path, _ = QFileDialog.getOpenFileName(
            self.view, "Importar Dados Principais (JSON)", "", "JSON Files (*.json)"
        )
        if not file_path: return

        reply = QMessageBox.question(
            self.view, "Confirmação de Importação",
            "A importação de dados principais (JSON) irá APAGAR TODOS os dados existentes na tabela 'Atas' "
            "e substituí-los pelos dados do arquivo JSON. Deseja continuar?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.No: return

        success, message = self.model.import_main_data_from_json(file_path)
        if success:
            QMessageBox.information(self.view, "Importação Concluída", message)
            self.check_db_status_and_load_data() # Re-verifica o status do DB e recarrega
        else:
            QMessageBox.critical(self.view, "Erro na Importação", message)

    def export_complementary_data_json(self):
        """Exporta os dados complementares (Status, Registros, Links) para um arquivo JSON."""
        file_path, _ = QFileDialog.getSaveFileName(
            self.view, "Exportar Dados Complementares (JSON)", "atas_complementares.json",
            "JSON Files (*.json);;All Files (*)"
        )
        if not file_path: return

        success, data = self.model.export_complementary_data_to_json()
        if success:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=4, ensure_ascii=False)
                QMessageBox.information(self.view, "Sucesso", f"Dados complementares exportados para:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self.view, "Erro", f"Não foi possível salvar o arquivo JSON:\n{e}")
        else:
            QMessageBox.critical(self.view, "Erro na Exportação", f"Ocorreu um erro ao exportar os dados complementares:\n{data}")

    def import_complementary_data_json(self):
        """Importa os dados complementares (Status, Registros, Links) de um arquivo JSON."""
        file_path, _ = QFileDialog.getOpenFileName(
            self.view, "Importar Dados Complementares (JSON)", "", "JSON Files (*.json)"
        )
        if not file_path: return

        reply = QMessageBox.question(
            self.view, "Confirmação de Importação",
            "A importação de dados complementares (JSON) irá APAGAR TODOS os dados existentes nas tabelas "
            "'Status', 'Registros' e 'Links' e substituí-los pelos dados do arquivo JSON. "
            "Certifique-se de que os dados principais (Atas) já foram importados, pois esta importação depende deles. Deseja continuar?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.No: return

        success, message = self.model.import_complementary_data_from_json(file_path)
        if success:
            QMessageBox.information(self.view, "Importação Concluída", message)
            self.check_db_status_and_load_data() # Re-verifica o status do DB e recarrega
        else:
            QMessageBox.critical(self.view, "Erro na Importação", message)

# ========================================= Métodos Auxiliares =========================================

    def _parse_date_string(self, date_string):
        if not date_string: return None
        try:
            return datetime.strptime(date_string, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return None

    def _create_dias_item(self, dias_restantes):
        item = QStandardItem(str(dias_restantes))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        font = item.font()
        font.setBold(True)
        item.setFont(font)
        if isinstance(dias_restantes, int):
            if dias_restantes < 0: item.setForeground(Qt.GlobalColor.red); item.setIcon(icon_manager.get_icon("head_skull"))
            elif dias_restantes <= 89: item.setForeground(QBrush(QColor("#FFA500"))); item.setIcon(icon_manager.get_icon("alert"))
            elif dias_restantes <= 179: item.setForeground(QBrush(QColor("#FFD700"))); item.setIcon(icon_manager.get_icon("mensagem"))
            else: item.setForeground(QBrush(QColor("#32CD32"))); item.setIcon(icon_manager.get_icon("aproved"))
        else: item.setForeground(QBrush(QColor("#AAAAAA"))); item.setIcon(icon_manager.get_icon("time"))
        return item

    def load_initial_data(self):
        """Carrega os dados e popula ambas as tabelas na inicialização."""
        if not self.model.db_initialized: # Não tenta carregar se o DB não está pronto
            return
        try:
            atas = self.model.get_all_atas()
            self.populate_table(atas)
            self.populate_previsualization_table()
        except Exception as e:
            QMessageBox.critical(self.view, "Erro", f"Não foi possível carregar os dados:\n{e}")

    def _get_status_style(self, status_text):
        """Retorna a cor e a fonte para um determinado status."""
        status_styles = {
            "SEÇÃO ATAS": (QColor("#FFFFFF"), QFont.Weight.Bold),
            "ATA GERADA": (QColor(230, 230, 150), QFont.Weight.Bold),
            "EMPRESA": (QColor(230, 230, 150), QFont.Weight.Bold),
            "SIGDEM": (QColor(230, 180, 100), QFont.Weight.Bold),
            "ASSINADO": (QColor(230, 180, 100), QFont.Weight.Bold),
            "PUBLICADO": (QColor(135, 206, 250), QFont.Weight.Bold),
            "PORTARIA": (QColor(230, 230, 150), QFont.Weight.Bold),
            "PORT. MARINHA": (QColor(135, 206, 250), QFont.Weight.Bold), # NOVO STATUS
            "ALERTA PRAZO": (QColor(255, 160, 160), QFont.Weight.Bold),
            "NOTA TÉCNICA": (QColor(255, 160, 160), QFont.Weight.Bold),
            "AGU": (QColor(255, 160, 160), QFont.Weight.Bold),
            "PRORROGADO": (QColor(135, 206, 250), QFont.Weight.Bold),
            "SIGAD" : (QColor(135, 206, 250), QFont.Weight.Bold),
            "PLANILHA" : (QColor(50, 205, 50), QFont.Weight.Bold) # Novo STATUS
        }
        color, weight = status_styles.get(status_text, (QColor("#FFFFFF"), QFont.Weight.Normal))
        return QBrush(color), weight

    def import_data(self):
        file_path, _ = QFileDialog.getOpenFileName(self.view, "Selecionar Planilha", "", "Planilhas Excel (*.xlsx)")
        if not file_path: return
        success, message = self.model.import_from_spreadsheet(file_path)
        if success:
            QMessageBox.information(self.view, "Importação Concluída", message)
            self.load_initial_data()
        else:
            QMessageBox.critical(self.view, "Erro na Importação", message)

    def delete_ata_by_parecer(self, parecer_value):
        reply = QMessageBox.question(self.view, "Confirmação", "Tem certeza que deseja excluir esta ata?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            if self.model.delete_ata(parecer_value):
                QMessageBox.information(self.view, "Sucesso", "Ata excluída com sucesso!")
                self.load_initial_data()
            else:
                QMessageBox.critical(self.view, "Erro", "Não foi possível excluir a ata.")

    def delete_selected_ata(self):
        selected_indexes = self.view.table_view.selectionModel().selectedRows()
        if not selected_indexes:
            QMessageBox.warning(self.view, "Nenhuma Seleção", "Selecione uma ou mais atas para excluir.")
            return
        reply = QMessageBox.question(self.view, "Confirmação", f"Excluir {len(selected_indexes)} ata(s)?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.No: return
        
        source_model = self.view.proxy_model.sourceModel()
        
        pareceres_to_delete = [source_model.item(self.view.proxy_model.mapToSource(idx).row(), 6).text() for idx in selected_indexes]
        
        for parecer in pareceres_to_delete:
            self.model.delete_ata(parecer)
        self.load_initial_data()

    def show_add_ata_dialog(self):
        from atas.view.add_ata_dialog import AddAtaDialog
        dialog = AddAtaDialog(self.view)
        dialog.ata_added.connect(self.add_new_ata)
        dialog.exec()

    def add_new_ata(self, ata_data):
        if self.model.add_ata(ata_data):
            QMessageBox.information(self.view, "Sucesso", "Ata adicionada!")
            self.load_initial_data()
        else:
            QMessageBox.critical(self.view, "Erro", "Não foi possível adicionar a ata.")

    def _create_centered_item(self, text):
        item = QStandardItem(str(text))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        return item

    def populate_table(self, atas: list):
        model = self.view.table_model
        model.clear()
        #headers = ["Dias", "Número", "Ano", "Empresa", "Ata", "Objeto", "Status"]
        headers = ["Dias", "Início", "Vencimento", "Pregão", "Ano", "Empresa", "Ata", "Objeto", "Status"]
        model.setHorizontalHeaderLabels(headers)
        today = date.today()
        for ata in atas:
            dias_restantes = "N/A"
            termino_formatado = "N/A"
            if ata.termino:
                if ata.termino == "2000-01-01":
                    dias_restantes = "AD"
                    termino_formatado = "01/01/2000"
                else:
                    termino_date = self._parse_date_string(ata.termino)
                    if termino_date:
                        dias_restantes = (termino_date - today).days
                        termino_formatado = termino_date.strftime("%d/%m/%Y")

            vigencia_inicio = "N/A"
            if getattr(ata, "celebracao", None):
                inicio_date = self._parse_date_string(ata.celebracao)
                if inicio_date:
                    vigencia_inicio = inicio_date.strftime("%d/%m/%Y")

            status_text = ata.status_info.status if ata.status_info else "SEÇÃO ATAS"
            status_item = self._create_centered_item(status_text)
            brush, weight = self._get_status_style(status_text)
            status_item.setForeground(brush)
            font = status_item.font()
            font.setWeight(weight)
            status_item.setFont(font)

            model.appendRow([
                self._create_dias_item(dias_restantes),         # 0 – Dias
                self._create_centered_item(vigencia_inicio),    # 1 – Vigência Início
                self._create_centered_item(termino_formatado),  # 2 – Vencimento
                self._create_centered_item(ata.numero),         # 3 – Pregão
                self._create_centered_item(ata.ano),            # 4 – Ano
                self._create_centered_item(ata.empresa),        # 5 – Empresa
                self._create_centered_item(ata.contrato_ata_parecer),  # 6 – Ata
                self._create_centered_item(ata.objeto),         # 7 – Objeto
                status_item                                      # 8 – Status
            ])

        # Configura as colunas
        header = self.view.table_view.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed);   header.resizeSection(0, 80)   # Dias
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed);   header.resizeSection(1, 95)  # Vigência Início
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed);   header.resizeSection(2, 95)   # Vencimento
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed);   header.resizeSection(3, 75)   # Pregão
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed);   header.resizeSection(4, 80)   # Ano
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)                                # Empresa
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed);   header.resizeSection(6, 170)  # Ata
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)                                # Objeto
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.Fixed);   header.resizeSection(8, 180)

    def show_details_on_double_click(self, index):
        source_index = self.view.proxy_model.mapToSource(index)
        row = source_index.row()
        source_model = self.view.proxy_model.sourceModel()
        
        # Coluna 6 é onde o número da Ata (Parecer) está guardado
        parecer_item = source_model.item(row, 6)
        if not parecer_item: return
        
        ata_data = self.model.get_ata_by_parecer(parecer_item.text())
        if ata_data:
            self.show_ata_details(ata_data)
        else:
            QMessageBox.warning(self.view, "Erro", f"A ata '{parecer_item.text()}' não foi encontrada no banco de dados!")

    def show_context_menu(self, position):
        index = self.view.table_view.indexAt(position)
        if not index.isValid(): return
        source_index = self.view.proxy_model.mapToSource(index)
        parecer = self.view.proxy_model.sourceModel().item(source_index.row(), 6).text()
        if not parecer: return
        
        menu = QMenu(self.view)
        ver_mais_action = menu.addAction(icon_manager.get_icon("init"), "Ver/Editar Detalhes")
        ver_mais_action.triggered.connect(lambda: self.show_details_on_double_click(index))
        menu.addSeparator()
        excluir_action = menu.addAction(icon_manager.get_icon("delete"), "Excluir esta ata")
        excluir_action.triggered.connect(lambda: self.delete_ata_by_parecer(parecer))
        menu.exec(self.view.table_view.mapToGlobal(position))

    def update_ata_from_dialog(self, dialog):
        """Pega TODOS os dados da janela (Geral e Fiscal) e salva no banco."""

        # 1. Coleta todos os dados da UI
        updated_data = dialog.get_updated_data()
        registros = [dialog.registro_list.item(i).text() for i in range(dialog.registro_list.count())]
        parecer_value = dialog.ata_data.contrato_ata_parecer

        try:
            # 2. Tenta salvar os dados principais (NUP, Setor, Objeto, etc.)
            success = self.model.update_ata(parecer_value, updated_data, registros)
            if not success:
                raise Exception("Ata principal não encontrada no banco de dados.")

            # 3. Tenta salvar os dados de fiscalização
            # (Agora que a importação está correta, isso deve funcionar)
            save_fiscalizacao_ata(self.model, parecer_value, dialog)

        except Exception as e:
            # 4. Se QUALQUER parte falhar, mostra o erro
            QMessageBox.critical(self.view, "Erro ao Salvar", f"Não foi possível atualizar a ata:\n{e}")
            return

        # 5. Se TUDO deu certo, recarrega a UI do diálogo e atualiza a tabela
        dialog.ata_data = self.model.get_ata_by_parecer(parecer_value)
        dialog.load_data() # Recarrega a UI do diálogo com os novos dados
        self.update_table_row(parecer_value)
        self.populate_previsualization_table()

        QMessageBox.information(dialog, "Sucesso", "Alterações salvas com sucesso!")

    def update_table_row(self, parecer_value: str):
        """
        Atualiza dinamicamente uma única linha na tabela principal após a edição.
        """
        try:
            # 1. Obter o modelo da tabela
            model = self.view.table_model
            
            # 2. Obter os novos dados atualizados do banco
            updated_ata = self.model.get_ata_by_parecer(parecer_value)
            if not updated_ata:
                print(f"Erro: Ata {parecer_value} não encontrada no banco. Recarregando tabela.")
                self.load_initial_data()
                return

            # 3. Encontrar a linha (row) que precisa ser atualizada
            # A coluna "Ata" (ID) é a [6]
            column_to_check = 6 
            row_to_update = -1
            
            for row in range(model.rowCount()):
                item = model.item(row, column_to_check)
                if item and item.text() == parecer_value:
                    row_to_update = row
                    break
            
            if row_to_update == -1:
                print(f"Erro: Ata {parecer_value} não encontrada na tabela. Recarregando tabela.")
                self.load_initial_data() # Segurança: se não achar, recarrega tudo
                return

            # 4. Preparar os dados formatados (lógica copiada do 'populate_table')
            
            # Col 0: Dias e Col 1: Vencimento
            dias_restantes = "N/A"
            termino_formatado = "N/A"
            
            if updated_ata.termino:
                if updated_ata.termino == "2000-01-01":
                    dias_restantes = "AD"
                    termino_formatado = "01/01/2000"
                else:
                    termino_date = self._parse_date_string(updated_ata.termino)
                    if termino_date:
                        dias_restantes = (termino_date - date.today()).days
                        termino_formatado = termino_date.strftime("%d/%m/%Y")
            
            vigencia_inicio = "N/A"
            if getattr(updated_ata, "celebracao", None):
                inicio_date = self._parse_date_string(updated_ata.celebracao)
                if inicio_date:
                    vigencia_inicio = inicio_date.strftime("%d/%m/%Y")
            
            # Col 8: Status
            status_text = updated_ata.status # O AtaData já trata o "SEÇÃO ATAS"
            status_item = self._create_centered_item(status_text)
            brush, weight = self._get_status_style(status_text)
            status_item.setForeground(brush)
            font = status_item.font()
            font.setWeight(weight)
            status_item.setFont(font)
            
            # 5. Atualizar CADA CÉLULA daquela linha
            model.setItem(row_to_update, 0, self._create_dias_item(dias_restantes))
            model.setItem(row_to_update, 1, self._create_centered_item(vigencia_inicio))
            model.setItem(row_to_update, 2, self._create_centered_item(termino_formatado))
            model.setItem(row_to_update, 3, self._create_centered_item(updated_ata.numero))
            model.setItem(row_to_update, 4, self._create_centered_item(updated_ata.ano))
            model.setItem(row_to_update, 5, self._create_centered_item(updated_ata.empresa))
            model.setItem(row_to_update, 6, self._create_centered_item(updated_ata.contrato_ata_parecer)) # A própria ID
            model.setItem(row_to_update, 7, self._create_centered_item(updated_ata.objeto))
            model.setItem(row_to_update, 8, status_item)
            
            # 6. Log de sucesso (o que você já estava vendo)
            print(f"✅ Linha da ata {parecer_value} atualizada na tabela.")
        
        except Exception as e:
            print(f"❌ Erro crítico ao atualizar linha da tabela: {e}. Recarregando tudo.")
            self.load_initial_data()

    def show_ata_details(self, ata_data):
        """Abre a janela de detalhes e conecta o sinal de atualização."""
        try:
            dialog = AtaDetailsDialog(ata_data, self.model, self.view)
            
            if hasattr(dialog, 'btn_sync_trello'):
                dialog.btn_sync_trello.clicked.connect(lambda: self.run_trello_sync_ata(dialog, ata_data))
            else:
                print("Aviso: Botão 'btn_sync_trello' não encontrado em AtaDetailsDialog (Ignorando...).")
                
            dialog.ata_updated.connect(lambda: self.update_ata_from_dialog(dialog))
            dialog.exec()
            
        except Exception as e:
            import traceback
            traceback.print_exc() # Mostra o erro verdadeiro no terminal
            QMessageBox.critical(self.view, "Erro ao abrir detalhes", f"Erro interno ao gerar a tela:\n{e}")

    def generate_excel_report(self):
        """
        Gera e salva uma planilha Excel com os dados das atas, usando formatação avançada
        e hyperlinks dinâmicos.
        """
        atas = self.model.get_all_atas()
        if not atas:
            QMessageBox.warning(self.view, "Nenhum Dado", "Não há atas para gerar a tabela.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self.view, "Salvar Planilha como...", "Relatorio_Atas_Administrativas.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return

        try:
            # --- PREPARAÇÃO DA PLANILHA ---
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Atas Administrativas"
            ano_atual = datetime.now().year
            data_atual_str = datetime.now().strftime("%d/%m/%Y")

            # --- ESTILOS ---
            title_font = Font(bold=True, size=14)
            subtitle_font = Font(bold=True, size=12)
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            link_font = Font(color="0000FF", underline="single") # Cor azul padrão de link

            # NOVO ESTILO para a coluna "DIAS P/ VENCIMENTO"
            days_to_expire_font = Font(bold=True, size=13, color="00B050") # Negrito, tamanho 13, cor verde

            # --- CABEÇALHO COM LOGOS ---
            try:
                path_icone = os.path.join('utils', 'icons', 'icone.ico')
                if os.path.exists(path_icone):
                    logo_esquerdo = Image(path_icone)
                    logo_esquerdo.height = 70; logo_esquerdo.width = 70
                    ws.add_image(logo_esquerdo, 'A1')
            except Exception as e:
                print(f"Aviso: Não foi possível carregar o ícone esquerdo: {e}")

            ws.merge_cells('B1:K3')
            ws['B1'].value = "CENTRO DE INTENDÊNCIA DA MARINHA EM BRASÍLIA\nDIVISÃO DE OBTENÇÃO"
            ws['B1'].font = title_font
            ws['B1'].alignment = center_align

            ws.merge_cells('A4:L4')
            ws['A4'].value = f"ACORDOS ADMINISTRATIVOS EM VIGOR {ano_atual}"
            ws['A4'].font = subtitle_font
            ws['A4'].alignment = center_align

            ws['L6'] = f"Data: {data_atual_str}"
            ws['L6'].font = Font(bold=True, italic=True)
            ws['L6'].alignment = Alignment(horizontal='center')

            # --- CABEÇALHOS DA TABELA ---
            headers = [
                "SETOR", "MODALIDADE", "N°", "ANO", "EMPRESA", "ATAS", "OBJETO",
                "CELEBRAÇÃO", "TERMO ADITIVO", "PORTARIA DE FISCALIZAÇÃO", "TÉRMINO", "DIAS P/ VENCIMENTO"
            ]
            ws.append(headers)
            header_row_num = ws.max_row
            for cell in ws[header_row_num]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align # Garante centralização dos cabeçalhos

            # --- DADOS DAS ATAS ---
            for ata in atas:

                # Pega os links e valores, garantindo que não sejam None
                parecer_val = ata.contrato_ata_parecer or ""
                has_parecer_link = bool(ata.links and ata.links.serie_ata_link)
                parecer_link = ata.links.serie_ata_link if has_parecer_link else ""

                termo_val = ata.termo_aditivo or ""
                has_ta_link = bool(ata.links and ata.links.ta_link)
                ta_link = ata.links.ta_link if has_ta_link else ""

                portaria_val = ata.portaria_fiscalizacao or ""
                has_portaria_link = bool(ata.links and ata.links.portaria_link)
                portaria_link = ata.links.portaria_link if has_portaria_link else ""

                data_celebracao_excel = self._parse_date_string(ata.celebracao)
                data_termino_excel = self._parse_date_string(ata.termino)

                # 1️⃣ Primeiro monta os dados da linha
                row_data = [
                    ata.setor,
                    ata.modalidade,
                    ata.numero,
                    ata.ano,
                    ata.empresa,
                    parecer_val,         # COLUNA F (6) - ATAS
                    ata.objeto,
                    data_celebracao_excel,
                    termo_val,           # COLUNA I (9)
                    portaria_val,        # COLUNA J (10)
                    data_termino_excel,  # COLUNA K (11)
                    f'=IF(ISBLANK(K{ws.max_row+1}), "N/A", K{ws.max_row+1}-TODAY())'
                ]

                # 2️⃣ Adiciona a linha na planilha
                ws.append(row_data)

                # 3️⃣ Pega o índice REAL da linha recém adicionada
                row_idx = ws.max_row
                current_row = ws[row_idx]

                # 4️⃣ Agora sim aplica os hyperlinks na célula certa
                if has_parecer_link:
                    cell = ws.cell(row=row_idx, column=6)  # F -> ATAS
                    cell.hyperlink = parecer_link
                    cell.font = link_font

                if has_ta_link:
                    cell = ws.cell(row=row_idx, column=9)  # I -> TERMO ADITIVO
                    cell.hyperlink = ta_link
                    cell.font = link_font

                if has_portaria_link:
                    cell = ws.cell(row=row_idx, column=10) # J -> PORTARIA
                    cell.hyperlink = portaria_link
                    cell.font = link_font

                # 5️⃣ Formatação da linha
                for cell in current_row:
                    cell.alignment = center_align  # Centraliza todas as células do corpo

                """if has_parecer_link: current_row[5].font = link_font
                if has_ta_link: current_row[8].font = link_font
                if has_portaria_link: current_row[9].font = link_font"""

                # Formato de data e dias
                current_row[7].number_format = 'DD/MM/YYYY'
                current_row[10].number_format = 'DD/MM/YYYY'

                # Aplica o NOVO ESTILO para a coluna "DIAS P/ VENCIMENTO" (índice 11)
                current_row[11].font = days_to_expire_font
                current_row[11].number_format = '0' # Mantém o formato numérico

            # --- AJUSTE FINAL DAS LARGURAS DAS COLUNAS ---
            ws.column_dimensions['A'].width = 12; ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 35; ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 45; ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15; ws.column_dimensions['J'].width = 25
            ws.column_dimensions['K'].width = 15; ws.column_dimensions['L'].width = 18 # Coluna L (Dias p/ Vencimento)

            workbook.save(file_path)
            QMessageBox.information(self.view, "Sucesso", f"Planilha salva com sucesso em:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self.view, "Erro ao Gerar Planilha", f"Ocorreu um erro ao gerar a planilha:\n{str(e)}")

    def generate_bi_export(self):
        """
        Gera uma planilha limpa (sem formatação visual complexa) focada apenas 
        nos dados necessários para alimentar um Dashboard de BI.
        """
        atas = self.model.get_all_atas()
        if not atas:
            QMessageBox.warning(self.view, "Nenhum Dado", "Não há atas para gerar a tabela de BI.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self.view, "Salvar Exportação para BI", "Dados_BI_Atas.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Dados BI"

            # Cabeçalhos focados em análise de dados
            headers = [
                'Setor', 'Modalidade', 'Empresa', 'Ata/Parecer', 'Objeto',
                'Status','Data de Celebração', 'Data de Término', 'Termo Aditivo', 
                'Valor Global', 'CNPJ', 'Link da Ata', 'Link da Portaria'
            ]
            ws.append(headers)

            # Estilo simples para o cabeçalho
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font

            # Inserção dos dados solicitados
            for ata in atas:
                # Extração segura dos links
                link_ata = ata.links.serie_ata_link if ata.links and getattr(ata.links, 'serie_ata_link', None) else "Sem link"
                link_portaria = ata.links.portaria_link if ata.links and getattr(ata.links, 'portaria_link', None) else "Sem link"

                # Formatação de datas para o Excel reconhecer (YYYY-MM-DD)
                # O BI lida melhor com formatos padronizados.
                celebracao = ata.celebracao if ata.celebracao and ata.celebracao != 'None' else ""
                termino = ata.termino if ata.termino and ata.termino != 'None' else ""

                ws.append([
                    ata.setor or "", 
                    ata.modalidade or "", 
                    ata.empresa or "",
                    ata.contrato_ata_parecer or "", 
                    ata.objeto or "",
                    ata.status_info.status or "", 
                    celebracao,
                    termino,
                    ata.termo_aditivo or "", 
                    ata.valor_global or "0", 
                    ata.cnpj or "", 
                    link_ata, 
                    link_portaria
                ])

            # Auto-ajuste simples das larguras para melhor leitura do usuário
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 50) # Limita a 50 para o objeto não ficar enorme

            workbook.save(file_path)
            QMessageBox.information(self.view, "Sucesso", f"Dados para BI exportados com sucesso para:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self.view, "Erro", f"Ocorreu um erro ao exportar os dados para BI: {e}")

    def generate_empty_template(self):
        """Gera uma planilha Excel vazia com os cabeçalhos necessários para a importação."""
        file_path, _ = QFileDialog.getSaveFileName(
            self.view, "Salvar Modelo de Planilha Vazia", "Modelo_Importacao_Atas.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Dados Atas"

            # Cabeçalhos exatos que o método de importação espera
            headers = [
                'SETOR', 'MODALIDADE', 'Nº/', 'ANO', 'EMPRESA', 
                'CONTRATO – ATA PARECER', 'OBJETO', 'CELEBRAÇÃO', 
                'TERMO ADITIVO', 'PORTARIA DE FISCALIZAÇÃO', 'TERMINO', 'OBSERVAÇÕES'
            ]
            ws.append(headers)

            # Estilo
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font

            workbook.save(file_path)
            QMessageBox.information(self.view, "Sucesso", f"Modelo de planilha vazia salvo em:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self.view, "Erro", f"Não foi possível gerar o modelo: {e}")

    def export_for_reimport(self):
        """Exporta todos os dados atuais para uma planilha no formato exato de importação."""
        atas = self.model.get_all_atas()
        if not atas:
            QMessageBox.warning(self.view, "Nenhum Dado", "Não há atas para exportar.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self.view, "Exportar Dados para Re-importação", "Backup_Atas_Importacao.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Dados Atas"

            headers = [
                'SETOR', 'MODALIDADE', 'Nº/', 'ANO', 'EMPRESA', 
                'CONTRATO – ATA PARECER', 'OBJETO', 'CELEBRAÇÃO', 
                'TERMO ADITIVO', 'PORTARIA DE FISCALIZAÇÃO', 'TERMINO', 'OBSERVAÇÕES'
            ]
            ws.append(headers)

            for ata in atas:
                ws.append([
                    ata.setor, ata.modalidade, ata.numero, ata.ano, ata.empresa,
                    ata.contrato_ata_parecer, ata.objeto, ata.celebracao,
                    ata.termo_aditivo, ata.portaria_fiscalizacao, ata.termino,
                    ata.observacoes
                ])

            workbook.save(file_path)
            QMessageBox.information(self.view, "Sucesso", f"Dados exportados com sucesso para:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self.view, "Erro", f"Ocorreu um erro ao exportar os dados: {e}")

# ================================================================ PRE-VISUALIZAÇÃO ==================================================
    def populate_previsualization_table(self):
        """Popula a tabela de pré-visualização com atas que não estão no status padrão."""
        model = self.view.preview_model
        model.clear()
        headers = ["Dias", "Ata", "Empresa", "Objeto", "Status"]
        model.setHorizontalHeaderLabels(headers)

        atas = self.model.get_atas_with_status_not_default()
        today = date.today()

        for ata in atas:
            dias_restantes = "N/A"
            if ata.termino:
                termino_date = self._parse_date_string(ata.termino)
                if termino_date:
                    dias_restantes = (termino_date - today).days

            dias_item = self._create_dias_item(dias_restantes)

            parecer_item = self._create_centered_item(ata.contrato_ata_parecer)
            # Guarda o parecer como identificador único
            parecer_item.setData(ata.contrato_ata_parecer, Qt.ItemDataRole.UserRole)

            status_text = ata.status_info.status if ata.status_info else "SEÇÃO ATAS"
            status_item = self._create_centered_item(status_text)
            brush, weight = self._get_status_style(status_text)
            status_item.setForeground(brush)
            font = status_item.font()
            font.setWeight(weight)
            status_item.setFont(font)

            model.appendRow([
                dias_item,
                parecer_item,
                self._create_centered_item(ata.empresa),
                self._create_centered_item(ata.objeto),
                status_item
            ])

        # Ajusta as colunas da tabela de pré-visualização
        header = self.view.preview_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed); header.resizeSection(0, 80)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed); header.resizeSection(1, 170)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed); header.resizeSection(4, 180)

    def show_details_on_preview_double_click(self, index):
        """Abre os detalhes da ata a partir da tabela de pré-visualização."""
        proxy_model = self.view.preview_proxy_model
        source_index = proxy_model.mapToSource(index)
        row = source_index.row()

        # O parecer está na coluna 1 e guardado no UserRole
        parecer_item = proxy_model.sourceModel().item(row, 1)
        if not parecer_item: return

        parecer_value = parecer_item.data(Qt.ItemDataRole.UserRole)
        if not parecer_value:
            parecer_value = parecer_item.text()

        ata_data = self.model.get_ata_by_parecer(parecer_value)

        if ata_data:
            self.show_ata_details(ata_data)
        else:
            QMessageBox.warning(self.view, "Aviso", f"A ata '{parecer_value}' não foi encontrada no banco de dados!")

# ================================================================ TRELLO SYNC ==================================================
    def run_trello_sync_ata(self, dialog, ata_data):
        """Inicia o processo de sincronização em segundo plano."""
        # 1. Feedback visual inicial
        dialog.btn_sync_trello.setEnabled(False)
        dialog.btn_sync_trello.setText("Sincronizando...")

        ui_data = dialog.get_updated_data()
        status_atual = ui_data.get('status', 'SEÇÃO ATAS')

        for key, value in ui_data.items():
            if hasattr(ata_data, key):
                setattr(ata_data, key, value)
        
        ata_data.status = status_atual

        trello_model = TrelloModel()
        trello_controller = TrelloIndividualController(trello_model)

        # Criamos uma classe worker similar à dos contratos
        self.trello_worker = TrelloSyncWorker(trello_controller, ata_data, status_atual)
        
        # Sobrescrevemos o método run do worker temporariamente para usar sync_ata
        self.trello_worker.run = lambda: self._execute_sync_ata(trello_controller, ata_data, status_atual)

        # 3. Conecta os sinais de término
        self.trello_worker.finished.connect(lambda success, msg: self._on_trello_sync_finished(success, msg, dialog))
        self.trello_worker.start()

    def _execute_sync_ata(self, controller, data, status):
        """Função auxiliar executada dentro da Thread."""
        try:
            success, message = controller.sync_ata(data, status)
            self.trello_worker.finished.emit(success, message)
        except Exception as e:
            self.trello_worker.finished.emit(False, str(e))

    def _on_trello_sync_finished(self, success, message, dialog):
        """Retorno após a conclusão da thread."""
        dialog.btn_sync_trello.setEnabled(True)
        dialog.btn_sync_trello.setText(" Sincronizar com Trello")
        
        if success:
            QMessageBox.information(dialog, "Trello", message)
        else:
            QMessageBox.warning(dialog, "Erro Trello", message)
