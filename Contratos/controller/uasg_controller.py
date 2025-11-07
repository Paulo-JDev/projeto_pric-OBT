from Contratos.view.main_window import MainWindow
from Contratos.model.uasg_model import UASGModel
from utils.utils import refresh_uasg_menu
from utils.icon_loader import icon_manager

from Contratos.view.details_dialog import DetailsDialog

from Contratos.controller.controller_table import populate_table, update_row_from_details
from Contratos.controller.mensagem_controller import MensagemController
from Contratos.controller.settings_controller import SettingsController
from Contratos.view.record_popup import RecordPopup
from Contratos.controller.manual_contract_controller import ManualContractController

from PyQt6.QtWidgets import QMessageBox, QMenu, QFileDialog, QApplication, QHeaderView
from PyQt6.QtGui import QStandardItem, QFont, QColor, QBrush
from PyQt6.QtCore import Qt, QSortFilterProxyModel, QRegularExpression
import requests
import sqlite3
import json
import os # Adicionado para os.path.expanduser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
import re
import shutil

class UASGController:
    def __init__(self, base_dir, parent_view=None): # Adicionamos parent_view por boas práticas
        from .dashboard_controller import DashboardController

        self.model = UASGModel(base_dir)
        self.view = MainWindow(self) # Cria sua "página"
        self.dashboard_controller = DashboardController(self.model, self.view)
        
        self.view.settings_button.clicked.connect(self.show_settings_dialog)
        self.manual_contract_ctrl = ManualContractController(self.model, self.view)

        self.loaded_uasgs = {}
        self.current_data = []
        self.filtered_data = []

        initial_mode = self.model.load_setting("data_mode", "Online")
        self.view.update_status_icon(initial_mode)
        self.view.update_clear_button_icon(initial_mode)

        if self.model.database_dir.exists():
            self.loaded_uasgs = self.model.load_saved_uasgs()
            print(f"📂 UASGs do módulo Contratos carregadas: {list(self.loaded_uasgs.keys())}")
        else:
            print("⚠ Diretório 'database' não encontrado. Nenhum dado de Contratos carregado.")

        self.load_saved_uasgs()
        self.populate_previsualization_table()

    def load_saved_uasgs(self):
        """Carrega as UASGs salvas e atualiza o menu."""
        self.loaded_uasgs = self.model.load_saved_uasgs()
        refresh_uasg_menu(self)  # Atualiza o menu após carregar as UASGs

    def add_uasg_to_menu(self, uasg):
        """Adiciona uma UASG ao menu suspenso."""
        action = self.view.menu_button.menu().addAction(f"UASG {uasg}", lambda: self.update_table(uasg))

    def fetch_and_create_table(self):
        """Busca os dados da UASG e atualiza o banco de dados."""
        uasg = self.view.uasg_input.text().strip()

        # Verificação se a UASG está vazia ou contém caracteres não numéricos
        if not uasg or not uasg.isdigit():
            QMessageBox.warning(self.view, "Entrada Inválida", "Por favor, insira um número UASG válido.")
            return

        try:
            if uasg in self.loaded_uasgs and self.loaded_uasgs[uasg]: # Verifica se há dados carregados
                QMessageBox.information(self.view, "Sucesso", f"UASG {uasg} carregada e salva com sucesso!")
                self.view.uasg_input.clear()
            else:
                # Se for nova, buscar e salvar
                data = self.model.fetch_uasg_data(uasg)
                if data is None:
                    QMessageBox.critical(self.view, "Erro de API", f"Erro ao buscar dados da UASG {uasg}.")
                    return

                self.model.save_uasg_data(uasg, data)
                self.loaded_uasgs[uasg] = data
                self.add_uasg_to_menu(uasg)
                #QMessageBox.information(self.view, "Sucesso", f"UASG {uasg} carregada e salva com sucesso!")
                self.view.tabs.setCurrentWidget(self.view.table_tab)
                self.populate_previsualization_table()
                
                # time.sleep(1) # Considere remover ou usar QStatusBar
                self.view.uasg_input.clear()

            self.loaded_uasgs = self.model.load_saved_uasgs()

            self.update_table(uasg)
            self.view.tabs.setCurrentWidget(self.view.table_tab)
        except requests.exceptions.RequestException as e: # Exceção mais específica
            QMessageBox.critical(self.view, "Erro de Rede", f"Erro de rede ao buscar UASG {uasg}: {str(e)}")
        except Exception as e: # Genérico para outros erros
            QMessageBox.critical(self.view, "Erro Inesperado", f"Erro inesperado ao processar UASG {uasg}: {str(e)}")

    def delete_uasg_data(self):
        """Deleta os dados da UASG informada e limpa a tabela se ela estiver em uso."""
        uasg_para_deletar = self.view.uasg_input.text().strip()
        if not uasg_para_deletar:
            QMessageBox.warning(self.view, "Entrada Inválida", "Por favor, insira um número UASG para deletar.")
            return

        if uasg_para_deletar not in self.loaded_uasgs:
            QMessageBox.warning(self.view, "Erro", f"Nenhum dado encontrado para a UASG {uasg_para_deletar}.")
            return

        info_label_text = self.view.uasg_info_label.text()
        
        # Extrai o código da UASG que está sendo exibida na tabela
        uasg_sendo_exibida = ""
        if "UASG: " in info_label_text:
            # Pega a parte depois de "UASG: " e antes do primeiro espaço
            partes = info_label_text.split(" ")
            if len(partes) > 1:
                uasg_sendo_exibida = partes[1]

        # Compara a UASG a ser deletada com a que está na tela
        if uasg_para_deletar == uasg_sendo_exibida:
            self.view.search_bar.clear()
            model = self.view.table.model()
            model.removeRows(0, model.rowCount())
            self.view.uasg_info_label.setText("UASG: -")
            print(f"Limpando a tabela, pois a UASG {uasg_para_deletar} está sendo visualizada.")
            
        
        # Procede com a deleção dos dados do banco de dados
        self.model.delete_uasg_data(uasg_para_deletar)
        self.loaded_uasgs.pop(uasg_para_deletar, None)
        
        # Atualiza o menu de UASGs e limpa o campo de input
        self.load_saved_uasgs()
        self.view.uasg_input.clear()
        
        QMessageBox.information(self.view, "Sucesso", f"UASG {uasg_para_deletar} deletada com sucesso.")

    def update_table(self, uasg):
        """Atualiza a tabela com os dados da UASG selecionada."""
        if uasg in self.loaded_uasgs:
            # Recarrega os dados do arquivo JSON para garantir que estão atualizados
            self.loaded_uasgs = self.model.load_saved_uasgs()
            
            # Atualiza os dados atuais com os dados recarregados
            if uasg in self.loaded_uasgs:
                self.current_data = self.loaded_uasgs[uasg]
                
                # Obter o nome resumido da UASG para mostrar no label
                nome_resumido = ""
                if self.current_data and len(self.current_data) > 0:
                    contrato = self.current_data[0]
                    nome_resumido = contrato.get("contratante", {}).get("orgao", {}).get("unidade_gestora", {}).get("nome_resumido", "")
                
                # Atualiza o label na interface
                self.view.uasg_info_label.setText(f"UASG: {uasg} - {nome_resumido}")
                
                # Popula a tabela com os dados usando a função do módulo controller_table
                populate_table(self, self.current_data)
                self.dashboard_controller.update_dashboard(self.current_data)
                print(f"✅ Tabela atualizada com os dados da UASG {uasg}.")
            else:
                # Limpa o label se não houver dados
                self.view.uasg_info_label.setText(f"UASG: -")
                print(f"⚠ UASG {uasg} não encontrada nos dados recarregados(especifico).")
        else:
            # Limpa o label se a UASG não for encontrada
            self.view.uasg_info_label.setText(f"UASG: -")
            print(f"⚠ UASG {uasg} não encontrada nos dados carregados(geral).")
        self.load_saved_uasgs()

    def clear_table(self):
        # Verifica se há dados carregados
        if not self.current_data or len(self.current_data) == 0:
            QMessageBox.warning(
                self.view,
                "Nenhuma UASG Carregada",
                "Não há contratos carregados no momento.\n\n"
                "Por favor, busque uma UASG antes de atualizar a tabela."
            )
            return
        
        # Pega o código da UASG do primeiro contrato carregado
        primeiro_contrato = self.current_data[0]
        uasg_code = primeiro_contrato.get("contratante", {}).get("orgao", {}).get("unidade_gestora", {}).get("codigo")
        
        if not uasg_code:
            QMessageBox.warning(
                self.view,
                "UASG Não Identificada",
                "Não foi possível identificar a UASG dos contratos carregados."
            )
            return
        
        try:
            print(f"🔄 Recarregando tabela da UASG {uasg_code}...")
            
            # ==================== ✅ USA O MÉTODO update_table ====================
            self.update_table(uasg_code)
            
            # Mensagem de sucesso
            QMessageBox.information(
                self.view, 
                "Atualização Concluída", 
                f"Tabela da UASG {uasg_code} recarregada com sucesso!\n\n"
                f"Total de contratos: {len(self.current_data)}"
            )
            
            print(f"✅ Tabela recarregada: {len(self.current_data)} contratos")
            
        except Exception as e:
            QMessageBox.critical(
                self.view,
                "Erro ao Atualizar",
                f"Erro ao recarregar a tabela:\n{str(e)}"
            )
            print(f"❌ Erro ao recarregar tabela: {e}")

    def show_context_menu(self, position):
        """
        ✅ ATUALIZADO: Exibe o menu de contexto ao clicar com o botão direito na tabela.
        Adiciona a opção de excluir apenas para contratos manuais, seguindo o modelo existente.
        """
        index = self.view.table.indexAt(position)
        if not index.isValid():
            return

        # Mapeia o índice filtrado para o índice do modelo base
        source_index = self.view.table.model().mapToSource(index)
        row = source_index.row()

        # Certifica-se de que a lista usada está correta
        data_source = self.current_data

        # Verifica se o índice é válido para evitar erro de "index out of range"
        if 0 <= row < len(data_source):
            contrato = data_source[row]
            menu = QMenu(self.view) # O parent do menu é a view principal

            # Adiciona o ícone "init" à ação "Ver Detalhes"
            details_action = menu.addAction(icon_manager.get_icon("init"), "Ver Detalhes") # ✅ Usei "detalhes" para o ícone, se "init" for um ícone válido, pode manter.
            details_action.triggered.connect(lambda: self.show_details_dialog(contrato.copy())) # ✅ Usando .copy() para segurança

            # ==================== ✅ ADICIONA OPÇÃO DE EXCLUIR SE FOR MANUAL ====================
            is_manual = contrato.get("manual", False)
            if is_manual:
                menu.addSeparator() # Adiciona um separador para organizar
                
                delete_action = menu.addAction(icon_manager.get_icon("delete"), "Excluir Contrato Manual")
                delete_action.triggered.connect(lambda: self._delete_manual_contract(contrato.copy(), row)) # ✅ Usando .copy() para segurança
            
            # Exibe o menu na posição do cursor
            menu.exec(self.view.table.viewport().mapToGlobal(position))
            # Não adicionamos deleteLater() ou WA_DeleteOnClose aqui,
            # confiando no comportamento padrão do QMenu.exec() e no gerenciamento de memória do Qt.

    def _delete_manual_contract(self, contrato_data, row):
        """
        ✅ NOVO MÉTODO: Exclui um contrato manual do banco de dados.
        Simplificado e focado na funcionalidade principal.
        """
        from Contratos.model.models import Contrato, StatusContrato, RegistroStatus, LinksContrato, Fiscalizacao
        
        contrato_id = contrato_data.get("id")
        contrato_numero = contrato_data.get("numero", "N/A")
        uasg_code = contrato_data.get("contratante", {}).get("orgao", {}).get("unidade_gestora", {}).get("codigo")
        
        # ==================== CONFIRMAÇÃO ====================
        reply = QMessageBox.question(
            self.view, # Usando self.view como parent para a QMessageBox
            "Confirmar Exclusão",
            f"⚠️ Tem certeza que deseja excluir o contrato manual?\n\n"
            f"Número: {contrato_numero}\n"
            f"ID: {contrato_id}\n"
            f"UASG: {uasg_code}\n\n"
            f"❌ Esta ação não pode ser desfeita!",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # ==================== EXCLUSÃO DO BANCO ====================
        db = self.model._get_db_session()
        
        try:
            # 1. Busca o contrato no banco
            contrato_db = db.query(Contrato).filter(Contrato.id == contrato_id).first()
            
            if not contrato_db:
                QMessageBox.warning(self.view, "Contrato Não Encontrado", f"O contrato {contrato_id} não foi encontrado no banco de dados.")
                return
            
            # 2. Deleta registros relacionados (garantindo a integridade)
            db.query(StatusContrato).filter(StatusContrato.contrato_id == contrato_id).delete(synchronize_session=False)
            db.query(RegistroStatus).filter(RegistroStatus.contrato_id == contrato_id).delete(synchronize_session=False)
            db.query(LinksContrato).filter(LinksContrato.contrato_id == contrato_id).delete(synchronize_session=False)
            db.query(Fiscalizacao).filter(Fiscalizacao.contrato_id == contrato_id).delete(synchronize_session=False)
            
            # 3. Deleta o contrato principal
            db.delete(contrato_db)
            db.commit()
            
            print(f"✅ Contrato manual {contrato_id} excluído com sucesso")
            
            # ==================== ATUALIZA A INTERFACE ====================
            # Remove da lista de dados atuais
            if row < len(self.current_data):
                self.current_data.pop(row)
            
            # Remove da tabela (usando o modelo da tabela)
            model = self.view.table.model()
            if model:
                model.removeRow(row)
            
            # Atualiza o dashboard (se houver um dashboard_controller)
            if hasattr(self, 'dashboard_controller'):
                self.dashboard_controller.update_dashboard(self.current_data)
            
            QMessageBox.information(
                self.view,
                "Exclusão Concluída",
                f"✅ Contrato manual {contrato_numero} excluído com sucesso!"
            )
            
        except Exception as e:
            db.rollback()
            QMessageBox.critical(self.view, "Erro ao Excluir", f"❌ Erro ao excluir o contrato manual:\n\n{str(e)}")
            print(f"❌ Erro ao excluir contrato manual: {e}")
        finally:
            db.close()

    def show_details_dialog(self, contrato):
        """Exibe o diálogo de detalhes do contrato."""
        details_dialog = DetailsDialog(contrato, self.model, self.view) # Passa self.model
        
        # Conectar o sinal data_saved ao método que atualiza a tabela
        details_dialog.data_saved.connect(self.update_table_from_details)
        
        details_dialog.exec()

    def update_table_from_details(self, details_info):
        """Atualiza a tabela quando os dados são salvos na DetailsDialog."""
        update_row_from_details(self, details_info)
        self.populate_previsualization_table()
    
    # =========================================== Método para abrir a janela de mensagens =================================================
    def show_msg_dialog(self):
        """
        Abre a janela de geração de mensagens para o contrato selecionado.
        """
        # Pega o índice da linha selecionada na tabela
        selected_indexes = self.view.table.selectionModel().selectedIndexes()
        
        if not selected_indexes:
            QMessageBox.warning(self.view, "Nenhum Contrato Selecionado", "Por favor, selecione um contrato na tabela antes de clicar em Mensagens.")
            return

        # Pega o índice real da fonte de dados
        source_index = self.view.table.model().mapToSource(selected_indexes[0])
        selected_row = source_index.row()
        
        # Pega os dados do contrato selecionado
        contract_data = self.current_data[selected_row]
        
        # Cria e exibe a nova janela de mensagens
        mensagem_controller = MensagemController(contract_data, self.model, parent=self.view)
        mensagem_controller.show()

    # =========================================== Método para abrir a janela de configurações =================================================
    def show_settings_dialog(self):
        """Abre a janela de configurações e conecta o sinal de mudança de modo."""
        settings_controller = SettingsController(self.model, self.view)
        settings_controller.mode_changed.connect(self.view.update_status_icon)
        settings_controller.mode_changed.connect(self.view.update_clear_button_icon)
        settings_controller.database_updated.connect(self.load_saved_uasgs)
        settings_controller.database_updated.connect(self._on_database_updated)
        settings_controller.show()

    # =========================================== Método para exportar e importar dados de status =================================================
    def export_status_data(self):
        """Exporta todos os dados de status para um arquivo JSON."""
        all_status_data = self.model.get_all_status_data()
        if not all_status_data:
            QMessageBox.information(self.view, "Exportar Status", "Não há dados de status para exportar.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self.view,
            "Salvar Dados de Status",
            "", # Diretório inicial
            "JSON Files (*.json);;All Files (*)"
        )

        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(all_status_data, f, ensure_ascii=False, indent=4)
                QMessageBox.information(self.view, "Exportar Status", f"Dados de status exportados com sucesso para:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self.view, "Erro ao Exportar", f"Não foi possível salvar o arquivo: {e}")

    def import_status_data(self):
        """Importa dados de status de um arquivo JSON."""
        file_path, _ = QFileDialog.getOpenFileName(
            self.view,
            "Abrir Dados de Status",
            "", # Diretório inicial
            "JSON Files (*.json);;All Files (*)"
        )

        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data_to_import = json.load(f)
                
                self.model.import_statuses(data_to_import)
                QMessageBox.information(self.view, "Importar Status", "Dados de status importados com sucesso!\nA tabela será atualizada.")
                self.load_saved_uasgs() # Recarrega UASGs e atualiza o menu
                # Força a atualização da tabela visível, se houver alguma UASG carregada
                current_uasg_text = self.view.uasg_info_label.text()
                if "UASG: " in current_uasg_text and current_uasg_text.split(" ")[1] != "-":
                    uasg_code = current_uasg_text.split(" ")[1]
                    self.update_table(uasg_code) # Atualiza a tabela com a UASG atual
                self.populate_previsualization_table()
            except json.JSONDecodeError:
                QMessageBox.critical(self.view, "Erro de Importação", "Arquivo JSON inválido ou corrompido.")
            except Exception as e:
                QMessageBox.critical(self.view, "Erro ao Importar", f"Não foi possível importar os dados: {e}")

# =========================================== Método para exportar a tabela para um arquivo Excel =================================================
    def export_table_to_excel(self):
        """
        Exporta os dados para um arquivo Excel (.xlsx) com formatação avançada,
        incluindo fórmulas dinâmicas e quebra de linha nos cabeçalhos.
        
        ✅ ATUALIZADO:
        - Usa objeto editado (se existir)
        - Adiciona bordas pretas em todas as células
        - Ícones nos cantos (A1 e K1)
        - Tamanhos de colunas em cm
        - Altura máxima de linhas: 0,80 cm
        - Coluna "DIAS P/ VENCIMENTO" em negrito, tamanho 13
        """
        if not self.current_data:
            QMessageBox.information(self.view, "Exportar Tabela", "Não há dados na tabela para exportar.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self.view,
            "Salvar Planilha como...",
            "Relatorio_Contratos.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            # ==================== IMPORTS NECESSÁRIOS ====================
            from openpyxl.styles import Border, Side
            
            # --- PREPARAÇÃO DA PLANILHA ---
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Acordos Administrativos"
            ano_atual = datetime.now().year
            data_atual_str = datetime.now().strftime("%d/%m/%Y")
            
            # ==================== DEFINE BORDA PRETA ====================
            thin_black_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # --- LÓGICA PARA INSERIR OS ÍCONES ---
            try:
                # ==================== ✅ ÍCONE ESQUERDO (A1) ====================
                path_icone = os.path.join('utils', 'icons', 'icone.ico')
                if os.path.exists(path_icone):
                    logo_esquerdo = Image(path_icone)
                    logo_esquerdo.height = 70
                    logo_esquerdo.width = 70
                    ws.add_image(logo_esquerdo, 'A1')
                else:
                    print(f"Aviso: Ícone não encontrado em {path_icone}")
                
                # ==================== ✅ ÍCONE DIREITO (K1) ====================
                path_acanto = os.path.join('utils', 'icons', 'acanto.png')
                if os.path.exists(path_acanto):
                    logo_direito = Image(path_acanto)
                    logo_direito.height = 70
                    logo_direito.width = 70
                    ws.add_image(logo_direito, 'K1')  # ✅ Mudado de L1 para K1
                else:
                    print(f"Aviso: Ícone não encontrado em {path_acanto}")
            except Exception as img_error:
                print(f"Ocorreu um erro ao carregar os logos: {img_error}")
                QMessageBox.warning(self.view, "Aviso", "Não foi possível carregar os logos na planilha.")
            
            # --- CABEÇALHO PRINCIPAL E TÍTULOS ---
            # ==================== ✅ MESCLA AJUSTADA (B1:J3) ====================
            ws.merge_cells('B1:J3')  # ✅ Mudado de B1:K3 para B1:J3
            ws['B1'].value = "CENTRO DE INTENDÊNCIA DA MARINHA EM BRASÍLIA\nDIVISÃO DE OBTENÇÃO"
            ws['B1'].font = Font(bold=True, size=14)
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # ==================== ✅ MESCLA AJUSTADA (A4:K4) ====================
            ws.merge_cells('A4:K4')  # ✅ Mudado de A4:L4 para A4:K4
            ws['A4'].value = f"ACORDOS ADMINISTRATIVOS EM VIGOR {ano_atual}"
            ws['A4'].font = Font(bold=True, size=12)
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
            
            # ==================== ✅ DATA MOVIDA PARA K6 ====================
            ref_date_cell = 'K6'  # ✅ Mudado de L6 para K6
            ws[ref_date_cell] = f"Data: {data_atual_str}"
            ws[ref_date_cell].font = Font(bold=True, italic=True)
            ws[ref_date_cell].alignment = Alignment(horizontal='center')
            
            # --- CABEÇALHOS DAS COLUNAS (com quebra de linha) ---
            headers = [
                "SETOR", "MODALIDADE", "N°/ANO", "EMPRESA", "CONTRATOS",
                "OBJETO", "CELEBRAÇÃO", "TERMO\nADITIVO", "PORTARIA DE\nFISCALIZAÇÃO", 
                "TÉRMINO", "DIAS P/\nVENCIMENTO"
            ]
            ws.append(headers)
            
            header_row = ws[7]
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            green_font = Font(color="00B050")
            link_font = Font(color="0000FF", underline="single")
            
            header_row_num = ws.max_row
            
            # ADICIONA BORDA NO CABEÇALHO
            for cell in ws[header_row_num]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_black_border
            
            # --- DADOS E FÓRMULAS ---
            center_alignment = Alignment(horizontal='center', vertical='center')
            green_font = Font(color="00B050")
            today = datetime.now().date()
            
            valid_contracts = []
            for uasg_list in self.loaded_uasgs.values():
                for contrato in uasg_list:
                    vigencia_fim_str = contrato.get("vigencia_fim")
                    if vigencia_fim_str:
                        try:
                            termino_date = datetime.strptime(vigencia_fim_str, "%Y-%m-%d").date()
                            dias_restantes = (termino_date - today).days
                            if dias_restantes >= 0:
                                contrato['dias_restantes'] = dias_restantes
                                valid_contracts.append(contrato)
                        except (ValueError, TypeError):
                            continue
            
            # Ordena por dias restantes (menor para maior)
            valid_contracts.sort(key=lambda x: x['dias_restantes'])
            
            # ==================== ✅ FONTE ESPECIAL PARA COLUNA "DIAS P/ VENCIMENTO" ====================
            dias_vencimento_font = Font(bold=True, size=13, color="00B050")  # ✅ Negrito, tamanho 13
            
            # --- PREENCHE DADOS ---
            for row_idx, contrato in enumerate(valid_contracts, start=header_row_num + 1):
                contrato_id = str(contrato.get("id"))
                
                # BUSCAR DADOS ADICIONAIS
                links_data = self.model.get_contract_links(contrato_id) or {}
                termo_aditivo_text = self._get_status_field_from_db(contrato_id, 'termo_aditivo_edit') or "XXX"
                
                # INSERIR DADOS E HIPERLINKS
                ws.cell(row=row_idx, column=1, value=contrato.get("contratante", {}).get("orgao", {}).get("unidade_gestora", {}).get("nome_resumido", "N/A"))
                ws.cell(row=row_idx, column=2, value=contrato.get("modalidade", "N/A"))
                ws.cell(row=row_idx, column=3, value=contrato.get("licitacao_numero", "N/A"))
                ws.cell(row=row_idx, column=4, value=contrato.get("fornecedor", {}).get("nome", ""))
                
                # Célula Contrato (Coluna E)
                cell_contrato = ws.cell(row=row_idx, column=5, value=contrato.get("numero", ""))
                link_pncp_espc = links_data.get("link_pncp_espc")
                if link_pncp_espc:
                    cell_contrato.hyperlink = link_pncp_espc
                    cell_contrato.font = link_font
                
                # OBJETO EDITADO (SE EXISTIR)
                objeto_editado = self._get_status_field_from_db(contrato_id, 'objeto_editado')
                if objeto_editado:
                    ws.cell(row=row_idx, column=6, value=objeto_editado)
                else:
                    ws.cell(row=row_idx, column=6, value=contrato.get("objeto", ""))
                
                # Célula Termo Aditivo (Coluna H)
                cell_ta = ws.cell(row=row_idx, column=8, value=termo_aditivo_text)
                link_ta = links_data.get("link_ta")
                if link_ta:
                    cell_ta.hyperlink = link_ta
                    cell_ta.font = link_font
                
                # Célula Portaria (Coluna I)
                portaria_text = self._get_status_field_from_db(contrato_id, 'portaria_edit') or "XXX"
                cell_portaria = ws.cell(row=row_idx, column=9, value=portaria_text)
                link_portaria = links_data.get("link_portaria")
                if link_portaria:
                    cell_portaria.hyperlink = link_portaria
                    cell_portaria.font = link_font
                
                # Datas
                data_celebracao_str = contrato.get("data_assinatura")
                if data_celebracao_str:
                    ws.cell(row=row_idx, column=7, value=datetime.strptime(data_celebracao_str, "%Y-%m-%d"))
                
                data_termino_str = contrato.get("vigencia_fim", "")
                if data_termino_str:
                    ws.cell(row=row_idx, column=10, value=datetime.strptime(data_termino_str, "%Y-%m-%d"))
                
                # ==================== ✅ COLUNA "DIAS P/ VENCIMENTO" COM FORMATAÇÃO ESPECIAL ====================
                cell_dias = ws.cell(row=row_idx, column=11, value=contrato['dias_restantes'])
                cell_dias.font = dias_vencimento_font  # ✅ Negrito, tamanho 13, verde
                cell_dias.number_format = '0'
                
                # Formatação da linha
                current_row = ws[row_idx]
                for cell in current_row:
                    cell.alignment = center_align
                    cell.border = thin_black_border  # ✅ Borda preta
                
                current_row[6].number_format = 'DD/MM/YYYY'
                current_row[9].number_format = 'DD/MM/YYYY'
                
                # ==================== ✅ ALTURA MÁXIMA DA LINHA: 0,80 CM ====================
                # 1 cm = 28.35 pontos (aproximadamente)
                # 0,80 cm = 22.68 pontos
                ws.row_dimensions[row_idx].height = 22.68  # ✅ 0,80 cm
            
            # ==================== ✅ LARGURAS DAS COLUNAS EM CM ====================
            # Conversão: 1 cm ≈ 3.78 unidades do Excel
            # Fórmula: largura_excel = (cm * 3.78) - 0.5
            
            column_widths_cm = {
                'A': 2.40,   # SETOR
                'B': 3.05,   # MODALIDADE
                'C': 2.32,   # N°/ANO
                'D': 7.10,   # EMPRESA
                'E': 3.00,   # CONTRATOS
                'F': 8.90,   # OBJETO
                'G': 2.90,   # CELEBRAÇÃO
                'H': 2.42,   # TERMO ADITIVO
                'I': 3.25,   # PORTARIA DE FISCALIZAÇÃO
                'J': 2.32,   # TÉRMINO
                'K': 3.00    # DIAS P/ VENCIMENTO
            }
            
            for col_letter, width_cm in column_widths_cm.items():
                # Converte cm para unidades do Excel
                width_excel = (width_cm * 5.67) - 0.5
                ws.column_dimensions[col_letter].width = width_excel
            
            workbook.save(file_path)
            QMessageBox.information(
                self.view, 
                "Exportação Concluída", 
                f"Planilha com todas as UASGs salva com sucesso em:\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self.view, 
                "Erro ao Exportar", 
                f"Ocorreu um erro ao gerar a planilha:\n{e}"
            )


    def _calculate_dias_restantes(self, vigencia_fim_str):
        from datetime import date, datetime # Mover imports para o topo do arquivo
        if vigencia_fim_str:
            try:
                vigencia_fim = datetime.strptime(vigencia_fim_str, "%Y-%m-%d").date()
                return (vigencia_fim - date.today()).days
            except ValueError:
                return "Erro Data"
        return "Sem Data"

    def _get_status_for_contrato(self, contrato_id):
        # Lógica similar à de controller_table.py para buscar o status do DB
        if contrato_id and self.model:
            try:
                conn = self.model._get_db_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT status FROM status_contratos WHERE contrato_id = ?", (contrato_id,))
                status_row = cursor.fetchone()
                conn.close()
                if status_row and status_row['status']:
                    return status_row['status']
            except sqlite3.Error:
                return "Erro DB"
        return "SEÇÃO CONTRATOS"
    
    def _get_status_field_from_db(self, contrato_id, field_name):
        """Busca um campo específico da tabela status_contratos para um contrato."""
        try:
            conn = self.model._get_db_connection()
            cursor = conn.cursor()
            cursor.execute(f"SELECT {field_name} FROM status_contratos WHERE contrato_id = ?", (contrato_id,))
            result = cursor.fetchone()
            conn.close()
            if result and result[field_name]:
                return result[field_name]
        except sqlite3.Error as e:
            print(f"Erro ao buscar campo '{field_name}' do DB: {e}")
        return None

    '''def set_pdf_download_folder(self):
        """Permite ao usuário definir a pasta de download para PDFs."""
        current_path = self.model.load_setting("pdf_download_path", os.path.join(os.path.expanduser("~"), "Downloads"))
        
        folder_path = QFileDialog.getExistingDirectory(
            self.view,
            "Selecionar Pasta para Salvar PDFs",
            current_path
        )
        if folder_path:
            self.model.save_setting("pdf_download_path", folder_path)
            QMessageBox.information(self.view, "Pasta Definida", f"Os PDFs serão salvos em:\n{folder_path}")'''
    
    def import_links_from_spreadsheet(self):
        """
        Abre uma planilha Excel e importa os links para os contratos correspondentes,
        aplicando filtros por UASG e vigência.
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self.view, "Selecionar Planilha de Links", "", "Planilhas Excel (*.xlsx)"
        )
        if not file_path:
            return

        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            header_row = [cell.value for cell in sheet[1]]
            expected_headers = ['link_contrato', 'termo_aditivo', 'portaria']
            if not all(h in header_row for h in expected_headers):
                QMessageBox.critical(self.view, "Erro de Formato", f"A planilha deve conter as colunas: {', '.join(expected_headers)}")
                return

            col_indices = {name: header_row.index(name) for name in expected_headers}
            
            sucesso_count = 0
            falha_count = 0
            
            # --- LÓGICA DE FILTRAGEM E MAPEAMENTO DOS CONTRATOS DO PROGRAMA ---
            program_contracts = {}
            today = datetime.now().date()
            for uasg_code, uasg_data in self.loaded_uasgs.items():
                for contract in uasg_data:
                    # Filtra por vigência: apenas contratos que não expiraram há mais de 40 dias
                    vigencia_fim_str = contract.get("vigencia_fim")
                    if vigencia_fim_str:
                        try:
                            termino_date = datetime.strptime(vigencia_fim_str, "%Y-%m-%d").date()
                            dias_restantes = (termino_date - today).days
                            if dias_restantes < -40:
                                continue # Pula este contrato
                        except (ValueError, TypeError):
                            continue # Pula se a data for inválida
                    
                    # Cria a chave composta (UASG, NUMERO/ANO)
                    numero_ano = self._normalize_contract_number(contract.get('numero', ''))
                    if numero_ano:
                        chave = (str(uasg_code), numero_ano)
                        program_contracts[chave] = contract

            print("--- Iniciando Importação de Links ---")
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                key_cell_obj = row[col_indices['link_contrato']]
                key_cell_value = key_cell_obj.value
                
                if not key_cell_value:
                    continue

                # Extrai a chave (UASG, NUMERO/ANO) da planilha
                chave_planilha = self._normalize_spreadsheet_key(key_cell_value)
                if not chave_planilha:
                    print(f"Linha {row_idx}: Formato inválido na coluna 'link_contrato' ('{key_cell_value}'), pulando.")
                    falha_count += 1
                    continue
                
                # Tenta encontrar o contrato correspondente usando a chave composta
                if chave_planilha in program_contracts:
                    contrato = program_contracts[chave_planilha]
                    contrato_id = str(contrato.get("id"))
                    
                    # Prepara os dados para salvar
                    link_data = {
                        'link_contrato': None,
                        'link_ta': None,
                        'link_portaria': None
                    }

                    if key_cell_obj.hyperlink:
                        link_data['link_contrato'] = key_cell_obj.hyperlink.target
                    else:
                        link_data['link_contrato'] = key_cell_value

                    # Lógica para Termo Aditivo e Portaria
                    cell_obj_ta = row[col_indices['termo_aditivo']]
                    termo_aditivo_text = cell_obj_ta.value
                    if termo_aditivo_text and str(termo_aditivo_text).strip().upper() != 'XXX':
                        if cell_obj_ta.hyperlink:
                            link_data['link_ta'] = cell_obj_ta.hyperlink.target
                        self.model.save_status_field(contrato_id, 'termo_aditivo_edit', str(termo_aditivo_text))

                    cell_obj_portaria = row[col_indices['portaria']]
                    portaria_text = cell_obj_portaria.value
                    if portaria_text and str(portaria_text).strip().upper() != 'XXX':
                        if cell_obj_portaria.hyperlink:
                            link_data['link_portaria'] = cell_obj_portaria.hyperlink.target
                        self.model.save_status_field(contrato_id, 'portaria_edit', str(portaria_text))
                    
                    self.model.save_contract_links(contrato_id, link_data)
                    sucesso_count += 1
                    print(f"Linha {row_idx}: Links para o contrato '{chave_planilha}' atualizados.")
                else:
                    print(f"Linha {row_idx}: Contrato '{chave_planilha}' não encontrado no programa (ou filtrado por vigência).")
                    falha_count += 1
            
            print("--- Importação Finalizada ---")
            QMessageBox.information(self.view, "Importação Concluída", f"Importação finalizada!\n\nSucessos: {sucesso_count}\nFalhas: {falha_count}")
            self.update_table(self.view.uasg_info_label.text().split(" ")[1])

        except Exception as e:
            QMessageBox.critical(self.view, "Erro ao Importar", f"Ocorreu um erro ao processar a planilha:\n{e}")

    def _normalize_spreadsheet_key(self, key_string):
        """Extrai (UASG, 'NUMERO/ANO') do formato da planilha 'UASG/ANO-NUM/00'."""
        match = re.search(r'(\d{5})/(\d{2,4})-(\d+)/', key_string)
        if match:
            uasg = match.group(1)
            year = match.group(2)
            number = int(match.group(3))

            # Mapeamento de UASG
            if uasg == '87000': uasg = '787000'
            elif uasg == '87010': uasg = '787010'
            
            # Converte ano de 2 dígitos para 4 dígitos
            if len(year) == 2:
                year = f"20{year}"
            
            numero_ano_formatado = f"{number:05d}/{year}" # Formato padronizado: 00001/2025
            return (uasg, numero_ano_formatado)
        return None

    # O método _normalize_contract_number permanece o mesmo.
    def _normalize_contract_number(self, contract_string):
        """Extrai (número, ano) do formato do programa 'NUMERO/ANO'."""
        if isinstance(contract_string, str) and '/' in contract_string:
            parts = contract_string.split('/')
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                number = int(parts[0])
                year = parts[1]
                return f"{number:05d}/{year}"
        return None

    
    # ======================================== Funções da Tabela-de-Pré-Visualização ==============================================================
    def populate_previsualization_table(self):
        """Busca os contratos com status diferente de 'SEÇÃO CONTRATOS' e popula a tabela de pré-visualização."""
        data = self.model.get_contracts_with_status_not_default()
        self.view.populate_preview_table(data)

    def show_records_popup(self, index):
        """Busca os registros de um contrato e os exibe em uma janela popup."""
        if not index.isValid():
            return

        proxy_model = self.view.preview_table.model()
        source_index = proxy_model.mapToSource(index)
        
        first_item = proxy_model.sourceModel().item(source_index.row(), 0)
        contrato_id = first_item.data(Qt.ItemDataRole.UserRole) if first_item else None

        if not contrato_id:
            print(f"Aviso: Não foi possível obter o ID do contrato para a linha {source_index.row()}.")
            return

        registros = []
        try:
            conn = self.model._get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT texto FROM registros_status WHERE contrato_id = ?", (contrato_id,))
            registros = [row['texto'] for row in cursor.fetchall()]
            conn.close()
        except sqlite3.Error as e:
            print(f"Erro ao buscar registros do DB: {e}")
            registros = ["Erro ao buscar registros."]

        # Passa o contrato_id para o popup e conecta o sinal ao novo método
        popup = RecordPopup(registros, contrato_id, self.view)
        popup.details_requested.connect(self.open_details_by_id)
        
        popup.move(self.view.cursor().pos())
        popup.exec()

    def open_details_by_id(self, contrato_id):
        """
        Encontra um contrato pelo seu ID em todas as UASGs carregadas
        e abre a janela de detalhes para ele.
        """
        contract_data_found = None
        # Procura o contrato em todos os dados carregados na aplicação
        for uasg_code, contracts_list in self.loaded_uasgs.items():
            for contract in contracts_list:
                # Compara o ID como string para garantir a correspondência
                if str(contract.get('id')) == contrato_id:
                    contract_data_found = contract
                    break
            if contract_data_found:
                break

        if contract_data_found:
            self.show_details_dialog(contract_data_found)
        else:
            QMessageBox.warning(self.view, "Erro", f"Não foi possível encontrar os dados completos para o contrato ID: {contrato_id}")

# ====================== Contratos Manuais ==========================
    def open_manual_contract_window(self):
            """Abre a janela de contratos manuais"""
            self.manual_contract_ctrl.open_manual_contract_window()

# ====================== Banco de Dados (settings) ==========================

    def _on_database_updated(self):
        """
        ✅ NOVO MÉTODO: Chamado quando o banco de dados é alterado nas configurações.
        
        Recarrega os dados e atualiza a interface automaticamente.
        """
        print("🔄 Banco de dados alterado, recarregando dados...")
        
        # 1. Recarrega as UASGs salvas do novo banco
        self.loaded_uasgs = self.model.load_saved_uasgs()
        
        # 2. Atualiza o menu de UASGs
        self.load_saved_uasgs()
        
        # 3. Atualiza a tabela de pré-visualização
        self.populate_previsualization_table()
        
        # 4. Se houver dados carregados na tabela principal, atualiza
        if self.current_data and len(self.current_data) > 0:
            primeiro_contrato = self.current_data[0]
            uasg_code = primeiro_contrato.get("contratante", {}).get("orgao", {}).get("unidade_gestora", {}).get("codigo")
            
            if uasg_code:
                # Verifica se a UASG ainda existe no novo banco
                if uasg_code in self.loaded_uasgs:
                    print(f"✅ Atualizando tabela da UASG {uasg_code}")
                    self.update_table(uasg_code)
                else:
                    # Se a UASG não existe mais, limpa a tabela
                    print(f"⚠️ UASG {uasg_code} não encontrada no novo banco, limpando tabela")
                    model = self.view.table.model()
                    if model:
                        model.removeRows(0, model.rowCount())
                    self.view.uasg_info_label.setText("UASG: -")
                    self.current_data = []
                    self.dashboard_controller.clear_dashboard()
        
        print("✅ Dados recarregados com sucesso!")
