# Contratos/controller/exp_imp_table_controller.py
# Controlador para exportação e importação em massa da tabela de contratos.

import os
import re
import json
import pandas as pd
from datetime import datetime, date
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import sqlite3

from Contratos.view.menus.table_options_dialog import TableOptionsDialog

class ExpImpTableController:
    """
    Controlador responsável por operações de Exportação e Importação em massa
    relacionadas à tabela de contratos (Excel, Links, BI).
    """
    def __init__(self, main_controller):
        """
        :param main_controller: Referência para o UASGController principal,
                                para acesso a model, view e dados carregados.
        """
        self.main_ctrl = main_controller
        self.model = main_controller.model
        self.view = main_controller.view

    # ==================== PROPERTIES (Proxy Dinâmico) ====================
    @property
    def current_data(self):
        """Pega sempre a lista atualizada do UASGController."""
        return self.main_ctrl.get_current_data()

    @property
    def loaded_uasgs(self):
        """Pega sempre o dicionário atualizado do UASGController."""
        return self.main_ctrl.get_loaded_uasgs()
    # =====================================================================

    def open_table_options_window(self):
        """Abre a mini janela com opções de tabela."""
        dialog = TableOptionsDialog(self.view)
        
        dialog.btn_export_excel.clicked.connect(self.export_table_to_excel)
        dialog.btn_import_links.clicked.connect(self.import_links_from_spreadsheet)
        dialog.btn_export_bi.clicked.connect(self.export_bi_data)
        
        dialog.exec()

    # =========================================================================
    # EXPORTAR TABELA (EXCEL)
    # =========================================================================
    def export_table_to_excel(self):
        """Exporta os dados da tabela para uma planilha Excel (.xlsx).
        ✅ ATUALIZADO:
        - Usa objeto editado (se existir)
        - Adiciona bordas pretas em todas as células
        - Ícones nos cantos (A1 e K1)
        - Tamanhos de colunas em cm
        - Altura máxima de linhas: 0,80 cm
        - Coluna "DIAS P/ VENCIMENTO" em negrito, tamanho 13
        """
        if not self.current_data:
            QMessageBox.information(self.view, "Exportar Tabela", "Não há dados na tabela para exportar.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self.view,
            "Salvar Planilha como...",
            "Relatorio_Contratos.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            # ==================== IMPORTS NECESSÁRIOS ====================
            from openpyxl.styles import Border, Side
            
            # --- PREPARAÇÃO DA PLANILHA ---
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Acordos Administrativos"
            ano_atual = datetime.now().year
            data_atual_str = datetime.now().strftime("%d/%m/%Y")
            
            # ==================== DEFINE BORDA PRETA ====================
            thin_black_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # --- LÓGICA PARA INSERIR OS ÍCONES ---
            try:
                # ==================== ✅ ÍCONE ESQUERDO (A1) ====================
                path_icone = os.path.join('utils', 'icons', 'icone.ico')
                if os.path.exists(path_icone):
                    logo_esquerdo = Image(path_icone)
                    logo_esquerdo.height = 70
                    logo_esquerdo.width = 70
                    ws.add_image(logo_esquerdo, 'A1')
                else:
                    print(f"Aviso: Ícone não encontrado em {path_icone}")
                
                # ==================== ✅ ÍCONE DIREITO (K1) ====================
                path_acanto = os.path.join('utils', 'icons', 'acanto.png')
                if os.path.exists(path_acanto):
                    logo_direito = Image(path_acanto)
                    logo_direito.height = 70
                    logo_direito.width = 70
                    ws.add_image(logo_direito, 'K1')  # ✅ Mudado de L1 para K1
                else:
                    print(f"Aviso: Ícone não encontrado em {path_acanto}")
            except Exception as img_error:
                print(f"Ocorreu um erro ao carregar os logos: {img_error}")
                QMessageBox.warning(self.view, "Aviso", "Não foi possível carregar os logos na planilha.")
            
            # --- CABEÇALHO PRINCIPAL E TÍTULOS ---
            # ==================== ✅ MESCLA AJUSTADA (B1:J3) ====================
            ws.merge_cells('B1:J3')  # ✅ Mudado de B1:K3 para B1:J3
            ws['B1'].value = "CENTRO DE INTENDÊNCIA DA MARINHA EM BRASÍLIA\nDIVISÃO DE OBTENÇÃO"
            ws['B1'].font = Font(bold=True, size=14)
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # ==================== ✅ MESCLA AJUSTADA (A4:K4) ====================
            ws.merge_cells('A4:K4')  # ✅ Mudado de A4:L4 para A4:K4
            ws['A4'].value = f"ACORDOS ADMINISTRATIVOS EM VIGOR {ano_atual}"
            ws['A4'].font = Font(bold=True, size=12)
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
            
            # ==================== ✅ DATA MOVIDA PARA K6 ====================
            ref_date_cell = 'K6'  # ✅ Mudado de L6 para K6
            ws[ref_date_cell] = f"Data: {data_atual_str}"
            ws[ref_date_cell].font = Font(bold=True, italic=True)
            ws[ref_date_cell].alignment = Alignment(horizontal='center')
            
            # --- CABEÇALHOS DAS COLUNAS (com quebra de linha) ---
            headers = [
                "SETOR", "MODALIDADE", "N°/ANO", "EMPRESA", "CONTRATOS",
                "OBJETO", "CELEBRAÇÃO", "TERMO\nADITIVO", "PORTARIA DE\nFISCALIZAÇÃO", 
                "TÉRMINO", "DIAS P/\nVENCIMENTO"
            ]
            ws.append(headers)
            
            header_row = ws[7]
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            green_font = Font(color="00B050")
            link_font = Font(color="0000FF", underline="single")
            
            header_row_num = ws.max_row
            
            # ADICIONA BORDA NO CABEÇALHO
            for cell in ws[header_row_num]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_black_border
            
            # --- DADOS E FÓRMULAS ---
            center_alignment = Alignment(horizontal='center', vertical='center')
            green_font = Font(color="00B050")
            today = datetime.now().date()
            
            valid_contracts = []
            for uasg_list in self.loaded_uasgs.values():
                for contrato in uasg_list:
                    vigencia_fim_str = contrato.get("vigencia_fim")
                    if vigencia_fim_str:
                        try:
                            termino_date = datetime.strptime(vigencia_fim_str, "%Y-%m-%d").date()
                            dias_restantes = (termino_date - today).days
                            if dias_restantes >= 0:
                                contrato['dias_restantes'] = dias_restantes
                                valid_contracts.append(contrato)
                        except (ValueError, TypeError):
                            continue
            
            # Ordena por dias restantes (menor para maior)
            valid_contracts.sort(key=lambda x: x['dias_restantes'])
            
            # ==================== ✅ FONTE ESPECIAL PARA COLUNA "DIAS P/ VENCIMENTO" ====================
            dias_vencimento_font = Font(bold=True, size=13, color="00B050")  # ✅ Negrito, tamanho 13
            
            # --- PREENCHE DADOS ---
            for row_idx, contrato in enumerate(valid_contracts, start=header_row_num + 1):
                contrato_id = str(contrato.get("id"))
                
                # BUSCAR DADOS ADICIONAIS
                links_data = self.model.get_contract_links(contrato_id) or {}
                termo_aditivo_text = self._get_status_field_from_db(contrato_id, 'termo_aditivo_edit') or "XXX"
                
                # INSERIR DADOS E HIPERLINKS
                ws.cell(row=row_idx, column=1, value=contrato.get("contratante", {}).get("orgao", {}).get("unidade_gestora", {}).get("nome_resumido", "N/A"))
                ws.cell(row=row_idx, column=2, value=contrato.get("modalidade", "N/A"))
                ws.cell(row=row_idx, column=3, value=contrato.get("licitacao_numero", "N/A"))
                ws.cell(row=row_idx, column=4, value=contrato.get("fornecedor", {}).get("nome", ""))
                
                # Célula Contrato (Coluna E)
                cell_contrato = ws.cell(row=row_idx, column=5, value=contrato.get("numero", ""))
                link_pncp_espc = links_data.get("link_pncp_espc")
                if link_pncp_espc:
                    cell_contrato.hyperlink = link_pncp_espc
                    cell_contrato.font = link_font
                
                # OBJETO EDITADO (SE EXISTIR)
                objeto_editado = self._get_status_field_from_db(contrato_id, 'objeto_editado')
                if objeto_editado:
                    ws.cell(row=row_idx, column=6, value=objeto_editado)
                else:
                    ws.cell(row=row_idx, column=6, value=contrato.get("objeto", ""))
                
                # Célula Termo Aditivo (Coluna H)
                cell_ta = ws.cell(row=row_idx, column=8, value=termo_aditivo_text)
                link_ta = links_data.get("link_ta")
                if link_ta:
                    cell_ta.hyperlink = link_ta
                    cell_ta.font = link_font
                
                # Célula Portaria (Coluna I)
                portaria_text = self._get_status_field_from_db(contrato_id, 'portaria_edit') or "XXX"
                cell_portaria = ws.cell(row=row_idx, column=9, value=portaria_text)
                link_portaria = links_data.get("link_portaria")
                if link_portaria:
                    cell_portaria.hyperlink = link_portaria
                    cell_portaria.font = link_font
                
                # Datas
                data_celebracao_str = contrato.get("data_assinatura")
                if data_celebracao_str:
                    ws.cell(row=row_idx, column=7, value=datetime.strptime(data_celebracao_str, "%Y-%m-%d"))
                
                data_termino_str = contrato.get("vigencia_fim", "")
                if data_termino_str:
                    ws.cell(row=row_idx, column=10, value=datetime.strptime(data_termino_str, "%Y-%m-%d"))
                
                # ==================== ✅ COLUNA "DIAS P/ VENCIMENTO" COM FORMATAÇÃO ESPECIAL ====================
                cell_dias = ws.cell(row=row_idx, column=11, value=contrato['dias_restantes'])
                cell_dias.font = dias_vencimento_font  # ✅ Negrito, tamanho 13, verde
                cell_dias.number_format = '0'
                
                # Formatação da linha
                current_row = ws[row_idx]
                for cell in current_row:
                    cell.alignment = center_align
                    cell.border = thin_black_border  # ✅ Borda preta
                
                current_row[6].number_format = 'DD/MM/YYYY'
                current_row[9].number_format = 'DD/MM/YYYY'
                
                # ==================== ✅ ALTURA MÁXIMA DA LINHA: 0,80 CM ====================
                # 1 cm = 28.35 pontos (aproximadamente)
                # 0,80 cm = 22.68 pontos
                ws.row_dimensions[row_idx].height = 22.68  # ✅ 0,80 cm
            
            # ==================== ✅ LARGURAS DAS COLUNAS EM CM ====================
            # Conversão: 1 cm ≈ 3.78 unidades do Excel
            # Fórmula: largura_excel = (cm * 3.78) - 0.5
            
            column_widths_cm = {
                'A': 2.40,   # SETOR
                'B': 3.05,   # MODALIDADE
                'C': 2.32,   # N°/ANO
                'D': 7.10,   # EMPRESA
                'E': 3.00,   # CONTRATOS
                'F': 8.90,   # OBJETO
                'G': 2.90,   # CELEBRAÇÃO
                'H': 2.42,   # TERMO ADITIVO
                'I': 3.25,   # PORTARIA DE FISCALIZAÇÃO
                'J': 2.32,   # TÉRMINO
                'K': 3.00    # DIAS P/ VENCIMENTO
            }
            
            for col_letter, width_cm in column_widths_cm.items():
                # Converte cm para unidades do Excel
                width_excel = (width_cm * 5.67) - 0.5
                ws.column_dimensions[col_letter].width = width_excel
            
            workbook.save(file_path)
            QMessageBox.information(
                self.view, 
                "Exportação Concluída", 
                f"Planilha com todas as UASGs salva com sucesso em:\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self.view, 
                "Erro ao Exportar", 
                f"Ocorreu um erro ao gerar a planilha:\n{e}"
            )

    # =========================================================================
    # IMPORTAR LINKS
    # =========================================================================
    def import_links_from_spreadsheet(self):
        """
        Abre uma planilha Excel e importa os links para os contratos correspondentes,
        aplicando filtros por UASG e vigência.
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self.view, "Selecionar Planilha de Links", "", "Planilhas Excel (*.xlsx)"
        )
        if not file_path:
            return

        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet = workbook.active
            
            header_row = [cell.value for cell in sheet[1]]
            expected_headers = ['link_contrato', 'termo_aditivo', 'portaria']
            if not all(h in header_row for h in expected_headers):
                QMessageBox.critical(self.view, "Erro de Formato", f"A planilha deve conter as colunas: {', '.join(expected_headers)}")
                return

            col_indices = {name: header_row.index(name) for name in expected_headers}
            
            sucesso_count = 0
            falha_count = 0
            
            # --- LÓGICA DE FILTRAGEM E MAPEAMENTO DOS CONTRATOS DO PROGRAMA ---
            program_contracts = {}
            today = datetime.now().date()
            for uasg_code, uasg_data in self.loaded_uasgs.items():
                for contract in uasg_data:
                    # Filtra por vigência: apenas contratos que não expiraram há mais de 40 dias
                    vigencia_fim_str = contract.get("vigencia_fim")
                    if vigencia_fim_str:
                        try:
                            termino_date = datetime.strptime(vigencia_fim_str, "%Y-%m-%d").date()
                            dias_restantes = (termino_date - today).days
                            if dias_restantes < -40:
                                continue # Pula este contrato
                        except (ValueError, TypeError):
                            continue # Pula se a data for inválida
                    
                    # Cria a chave composta (UASG, NUMERO/ANO)
                    numero_ano = self._normalize_contract_number(contract.get('numero', ''))
                    if numero_ano:
                        chave = (str(uasg_code), numero_ano)
                        program_contracts[chave] = contract

            print("--- Iniciando Importação de Links ---")
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                key_cell_obj = row[col_indices['link_contrato']]
                key_cell_value = key_cell_obj.value
                
                if not key_cell_value:
                    continue

                # Extrai a chave (UASG, NUMERO/ANO) da planilha
                chave_planilha = self._normalize_spreadsheet_key(key_cell_value)
                if not chave_planilha:
                    print(f"Linha {row_idx}: Formato inválido na coluna 'link_contrato' ('{key_cell_value}'), pulando.")
                    falha_count += 1
                    continue
                
                # Tenta encontrar o contrato correspondente usando a chave composta
                if chave_planilha in program_contracts:
                    contrato = program_contracts[chave_planilha]
                    contrato_id = str(contrato.get("id"))
                    
                    # Prepara os dados para salvar
                    link_data = {
                        'link_contrato': None,
                        'link_ta': None,
                        'link_portaria': None
                    }

                    if key_cell_obj.hyperlink:
                        link_data['link_contrato'] = key_cell_obj.hyperlink.target
                    else:
                        link_data['link_contrato'] = key_cell_value

                    # Lógica para Termo Aditivo e Portaria
                    cell_obj_ta = row[col_indices['termo_aditivo']]
                    termo_aditivo_text = cell_obj_ta.value
                    if termo_aditivo_text and str(termo_aditivo_text).strip().upper() != 'XXX':
                        if cell_obj_ta.hyperlink:
                            link_data['link_ta'] = cell_obj_ta.hyperlink.target
                        self.model.save_status_field(contrato_id, 'termo_aditivo_edit', str(termo_aditivo_text))

                    cell_obj_portaria = row[col_indices['portaria']]
                    portaria_text = cell_obj_portaria.value
                    if portaria_text and str(portaria_text).strip().upper() != 'XXX':
                        if cell_obj_portaria.hyperlink:
                            link_data['link_portaria'] = cell_obj_portaria.hyperlink.target
                        self.model.save_status_field(contrato_id, 'portaria_edit', str(portaria_text))
                    
                    self.model.save_contract_links(contrato_id, link_data)
                    sucesso_count += 1
                    print(f"Linha {row_idx}: Links para o contrato '{chave_planilha}' atualizados.")
                else:
                    print(f"Linha {row_idx}: Contrato '{chave_planilha}' não encontrado no programa (ou filtrado por vigência).")
                    falha_count += 1
            
            print("--- Importação Finalizada ---")
            QMessageBox.information(self.view, "Importação Concluída", f"Importação finalizada!\n\nSucessos: {sucesso_count}\nFalhas: {falha_count}")
            self.update_table(self.view.uasg_info_label.text().split(" ")[1])

        except Exception as e:
            QMessageBox.critical(self.view, "Erro ao Importar", f"Ocorreu um erro ao processar a planilha:\n{e}")

    # =========================================================================
    # EXPORTAR DADOS BI
    # =========================================================================
    def export_bi_data(self):
        """
        Exporta dados para Excel (BI), incluindo:
        - UASG (código e nome)
        - Vigência início e fim
        - Duas abas: Contratos Vigentes e Contratos Mortos (com data no nome da aba)
        """
        import pandas as pd
        from datetime import datetime, date

        # 1. Escolher onde salvar
        hoje_str = datetime.now().strftime('%d-%m-%Y')
        filename, _ = QFileDialog.getSaveFileName(
            self.view,
            "Salvar Relatório BI",
            f"Relatorio_BI_{hoje_str}.xlsx",
            "Arquivos Excel (*.xlsx)"
        )
        if not filename:
            return

        conn = None
        try:
            # 2. Conectar ao Banco
            if hasattr(self, 'model') and hasattr(self.model, '_get_db_connection'):
                conn = self.model._get_db_connection()
            else:
                root_dir = os.getcwd()
                db_path = os.path.join(root_dir, "database", "gerenciador_uasg.db")
                if not os.path.exists(db_path):
                    db_path = os.path.join(root_dir, "..", "database", "gerenciador_uasg.db")
                if not os.path.exists(db_path):
                    raise FileNotFoundError("Banco de dados não encontrado.")
                conn = sqlite3.connect(db_path)

            # 3. Query SQL (agora trazendo UASG nome + vigências)
            query = """
            SELECT 
                c.numero AS numero_contrato,
                c.uasg_code AS uasg,
                c.contratante_orgao_unidade_gestora_nome_resumido AS uasg_nome,
                c.valor_global,
                COALESCE(s.objeto_editado, c.objeto) AS objeto_final,
                COALESCE(s.status, 'SEÇÃO CONTRATOS') AS status_atual,
                c.tipo,
                c.modalidade,
                s.radio_options_json,
                c.vigencia_inicio,
                c.vigencia_fim
            FROM contratos c
            LEFT JOIN status_contratos s ON c.id = s.contrato_id
            """

            # 4. Carregar no Pandas
            df = pd.read_sql_query(query, conn)
            if df.empty:
                QMessageBox.warning(self.view, "Aviso", "A tabela está vazia.")
                return

            # 5. Material/Serviço (lê do JSON)
            def get_material_servico(row):
                json_str = row.get('radio_options_json')
                if not json_str:
                    return "Definir"
                try:
                    data = json.loads(json_str)
                    valor = data.get("Material/Serviço:") or data.get("Material/Serviço")
                    if not valor or valor == "Não selecionado":
                        return "Definir"
                    return valor
                except:
                    return "Definir"

            df['Material/Serviço'] = df.apply(get_material_servico, axis=1)

            # 6. Formatação do Número do Contrato
            def formatar_numero(row):
                try:
                    uasg = str(row['uasg'])[-5:] if row.get('uasg') else "00000"
                    numero_raw = str(row.get('numero_contrato', ''))
                    if "/" in numero_raw:
                        parts = numero_raw.split('/')
                        num_parte = parts[0].lstrip('0') or "0"
                        ano_parte = parts[1][-2:] if len(parts[1]) >= 2 else parts[1]
                        return f"{uasg}/{ano_parte}-{num_parte}/00"
                    return numero_raw
                except:
                    return str(row.get('numero_contrato', ''))

            df['Número Formatado'] = df.apply(formatar_numero, axis=1)

            # 7. Converter vigências para data (pandas)
            # Aceita "YYYY-MM-DD"; se vier vazio vira NaT
            df['vigencia_inicio_dt'] = pd.to_datetime(df['vigencia_inicio'], errors='coerce').dt.date
            df['vigencia_fim_dt'] = pd.to_datetime(df['vigencia_fim'], errors='coerce').dt.date

            # 7.1 Criar colunas formatadas para exportação (dd/mm/aaaa)
            df['vigencia_inicio_export'] = pd.to_datetime(df['vigencia_inicio'], errors='coerce').dt.strftime('%d/%m/%Y')
            df['vigencia_fim_export'] = pd.to_datetime(df['vigencia_fim'], errors='coerce').dt.strftime('%d/%m/%Y')

            # Se preferir célula vazia ao invés de "NaT"
            df['vigencia_inicio_export'] = df['vigencia_inicio_export'].fillna("")
            df['vigencia_fim_export'] = df['vigencia_fim_export'].fillna("")

            hoje = date.today()

            # Regra:
            # - Vigente: vigencia_fim >= hoje E (vigencia_inicio é nula OU <= hoje)
            # - Morto: vigencia_fim < hoje
            # - Sem vigencia_fim: vamos considerar como "vigente" (para não sumir do relatório)
            def is_vigente(row):
                ini = row['vigencia_inicio_dt']
                fim = row['vigencia_fim_dt']
                if fim is None or pd.isna(fim):
                    return True
                if ini is None or pd.isna(ini):
                    return fim >= hoje
                return (ini <= hoje) and (fim >= hoje)

            df['__vigente__'] = df.apply(is_vigente, axis=1)

            df_vigentes = df[df['__vigente__'] == True].copy()
            df_mortos = df[(df['vigencia_fim_dt'].notna()) & (df['vigencia_fim_dt'] < hoje)].copy()

            # 8. Selecionar e renomear colunas finais
            colunas_finais = [
                'Número Formatado',
                'uasg',
                'uasg_nome',
                'Material/Serviço',
                'valor_global',
                'objeto_final',
                'status_atual',
                'tipo',
                'modalidade',
                'vigencia_inicio_export',
                'vigencia_fim_export'
            ]

            cols_to_export = [c for c in colunas_finais if c in df.columns]

            def montar_df_export(dfx):
                out = dfx[cols_to_export].copy()
                rename_map = {
                    'Número Formatado': 'Número',
                    'uasg': 'UASG Código',
                    'uasg_nome': 'UASG Nome',
                    'valor_global': 'Valor Global',
                    'objeto_final': 'Objeto',
                    'status_atual': 'Status',
                    'tipo': 'Tipo',
                    'modalidade': 'Modalidade',
                    'vigencia_inicio_export': 'Vigência Início',
                    'vigencia_fim_export': 'Vigência Fim',
                }
                out.rename(columns=rename_map, inplace=True)
                return out

            df_vigentes_final = montar_df_export(df_vigentes)
            df_mortos_final = montar_df_export(df_mortos)

            # 9. Gerar Excel com 2 abas
            sheet_vivos = "Contratos Vigentes"
            sheet_mortos = f"Contratos Mortos {hoje_str}"  # <= 31 chars OK

            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df_vigentes_final.to_excel(writer, sheet_name=sheet_vivos, index=False)
                df_mortos_final.to_excel(writer, sheet_name=sheet_mortos, index=False)

            QMessageBox.information(
                self.view,
                "Sucesso",
                f"Relatório BI exportado!\n\n"
                f"Aba 1: {sheet_vivos} ({len(df_vigentes_final)} contratos)\n"
                f"Aba 2: {sheet_mortos} ({len(df_mortos_final)} contratos)\n\n"
                f"{filename}"
            )

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            QMessageBox.critical(self.view, "Erro", f"Erro ao exportar: {str(e)}")

        finally:
            if conn:
                conn.close()
    # =========================================================================
    # MÉTODOS AUXILIARES (PRIVADOS)
    # =========================================================================

    def _get_status_field_from_db(self, contrato_id, field_name):
        """Busca um campo específico da tabela status_contratos para um contrato."""
        try:
            conn = self.model._get_db_connection()
            cursor = conn.cursor()
            cursor.execute(f"SELECT {field_name} FROM status_contratos WHERE contrato_id = ?", (contrato_id,))
            result = cursor.fetchone()
            conn.close()
            if result and result[field_name]:
                return result[field_name]
        except sqlite3.Error as e:
            print(f"Erro ao buscar campo '{field_name}' do DB: {e}")
        return None

    def _normalize_spreadsheet_key(self, key_string):
        """Extrai (UASG, 'NUMERO/ANO') do formato da planilha 'UASG/ANO-NUM/00'."""
        match = re.search(r'(\d{5})/(\d{2,4})-(\d+)/', key_string)
        if match:
            uasg = match.group(1)
            year = match.group(2)
            number = int(match.group(3))

            # Mapeamento de UASG
            if uasg == '87000': uasg = '787000'
            elif uasg == '87010': uasg = '787010'
            
            # Converte ano de 2 dígitos para 4 dígitos
            if len(year) == 2:
                year = f"20{year}"
            
            numero_ano_formatado = f"{number:05d}/{year}" # Formato padronizado: 00001/2025
            return (uasg, numero_ano_formatado)
        return None

    # O método _normalize_contract_number permanece o mesmo.
    def _normalize_contract_number(self, contract_string):
        """Extrai (número, ano) do formato do programa 'NUMERO/ANO'."""
        if isinstance(contract_string, str) and '/' in contract_string:
            parts = contract_string.split('/')
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                number = int(parts[0])
                year = parts[1]
                return f"{number:05d}/{year}"
        return None
