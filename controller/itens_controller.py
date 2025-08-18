# controller/itens_controller.py

from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

class ItensController:
    def __init__(self, model, parent_view=None):
        self.model = model
        self.parent_view = parent_view

    @staticmethod
    def _to_float(value_str):
        if not isinstance(value_str, str): return 0.0
        return float(value_str.replace('.', '').replace(',', '.'))

    def generate_report_to_excel(self, contract_data):
        """
        Orquestra a busca de dados, o processamento e a criação do relatório Excel dos itens.
        """
        contrato_id = contract_data.get("id")

        itens, error_itens = self.model.get_sub_data_for_contract(contrato_id, "itens")

        if error_itens or not itens:
            QMessageBox.critical(self.parent_view, "Erro de Dados", "Não foi possível buscar os itens para gerar o relatório.")
            return

        # Pede ao usuário onde salvar o arquivo
        file_path, _ = QFileDialog.getSaveFileName(
            self.parent_view,
            "Salvar Relatório de Itens",
            f"Relatorio_Itens_Contrato_{contract_data.get('numero', '').replace('/', '-')}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            self._create_excel_file(file_path, itens)
            QMessageBox.information(self.parent_view, "Sucesso", f"Relatório de itens salvo com sucesso em:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self.parent_view, "Erro ao Gerar Excel", f"Ocorreu um erro ao criar a planilha:\n{str(e)}")

    def _create_excel_file(self, file_path, itens_data):
        """Cria o arquivo .xlsx com a aba de resumo e as abas detalhadas para o histórico de cada item."""
        workbook = Workbook()

        body_font = Font(color="FFFFFF") # Letra branca
        body_fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid") # Fundo preto/cinza escuro
        
        # --- 1. ABA DE RESUMO GERAL ---
        ws_resumo = workbook.active
        ws_resumo.title = "Resumo Geral dos Itens"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        
        headers_resumo = [
            "ID do Item", "Tipo", "Grupo", "CATMAT/SER", "Descrição Complementar",
            "Quantidade", "Valor Unitário", "Valor Total"
        ]
        ws_resumo.append(headers_resumo)
        for cell in ws_resumo[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for item in itens_data:
            ws_resumo.append([
                item.get('id'), item.get('tipo_id'), item.get('grupo_id'),
                item.get('catmatseritem_id'), item.get('descricao_complementar'),
                self._to_float(item.get('quantidade', '0')),
                self._to_float(item.get('valorunitario', '0,00')),
                self._to_float(item.get('valortotal', '0,00'))
            ])

        for row in ws_resumo.iter_rows(min_row=2, max_row=ws_resumo.max_row):
            for cell in row: # Itera sobre cada célula da linha
                cell.font = body_font
                cell.fill = body_fill
            row[6].number_format = '"R$" #,##0.00'
            row[7].number_format = '"R$" #,##0.00'

        for i, col in enumerate(ws_resumo.columns):
            ws_resumo.column_dimensions[get_column_letter(i + 1)].auto_size = True

        # --- 2. ABAS DETALHADAS PARA O HISTÓRICO DE CADA ITEM ---
        for item in itens_data:
            sheet_title = f"Histórico Item {item.get('id')}"[:31]
            ws_detalhe = workbook.create_sheet(title=sheet_title)

            headers_detalhe = ["Tipo do Histórico", "Data do Termo", "Quantidade", "Periodicidade", "Valor Unitário", "Valor Total"]
            ws_detalhe.append(headers_detalhe)
            for cell in ws_detalhe[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            for historico in item.get('historico_item', []):
                ws_detalhe.append([
                    historico.get('tipo_historico'),
                    datetime.strptime(historico.get('data_termo'), "%Y-%m-%d").date() if historico.get('data_termo') else "N/A",
                    self._to_float(historico.get('quantidade', '0')),
                    historico.get('periodicidade'),
                    self._to_float(historico.get('valor_unitario', '0,00')),
                    self._to_float(historico.get('valor_total', '0,00'))
                ])

            for row in ws_detalhe.iter_rows(min_row=2, max_row=ws_detalhe.max_row):
                for cell in row: # Itera sobre cada célula da linha
                    cell.font = body_font
                    cell.fill = body_fill
                row[1].alignment = center_align
                row[1].number_format = 'DD/MM/YYYY'
                row[4].number_format = '"R$" #,##0.00'
                row[5].number_format = '"R$" #,##0.00'

            for i, col in enumerate(ws_detalhe.columns):
                ws_detalhe.column_dimensions[get_column_letter(i + 1)].auto_size = True

        workbook.save(file_path)
