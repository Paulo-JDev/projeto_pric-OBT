from view.main_window import MainWindow
from model.uasg_model import UASGModel
from utils.utils import refresh_uasg_menu
from utils.icon_loader import icon_manager

from view.details_dialog import DetailsDialog

from controller.controller_table import populate_table, update_row_from_details
from controller.mensagem_controller import MensagemController
from controller.settings_controller import SettingsController

from PyQt6.QtWidgets import QMessageBox, QMenu, QFileDialog, QApplication, QHeaderView
from PyQt6.QtGui import QStandardItem, QFont, QColor, QBrush
from PyQt6.QtCore import Qt, QSortFilterProxyModel, QRegularExpression
import requests
import sqlite3
import json
import os # Adicionado para os.path.expanduser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime

class UASGController:
    def __init__(self, base_dir):
        from controller.dashboard_controller import DashboardController

        self.model = UASGModel(base_dir)
        self.view = MainWindow(self)
        self.dashboard_controller = DashboardController(self.model, self.view)
        self.view.settings_button.clicked.connect(self.show_settings_dialog)

        # Dados carregados
        self.loaded_uasgs = {}
        self.current_data = []
        self.filtered_data = []

        initial_mode = self.model.load_setting("data_mode", "Online")
        self.view.update_status_icon(initial_mode)
        self.view.update_clear_button_icon(initial_mode)

        # Verifica se o diretório database existe
        if self.model.database_dir.exists():
            self.loaded_uasgs = self.model.load_saved_uasgs()
            print(f"📂 UASGs carregadas: {list(self.loaded_uasgs.keys())}")
        else:
            print("⚠ Diretório 'database' não encontrado. Nenhum dado carregado.")

        # Carrega as UASGs salvas e atualiza o menu
        self.load_saved_uasgs()
        self.populate_previsualization_table()


    def run(self):
        """Executa a interface principal."""
        self.view.show()

    def load_saved_uasgs(self):
        """Carrega as UASGs salvas e atualiza o menu."""
        self.loaded_uasgs = self.model.load_saved_uasgs()
        refresh_uasg_menu(self)  # Atualiza o menu após carregar as UASGs

    def add_uasg_to_menu(self, uasg):
        """Adiciona uma UASG ao menu suspenso."""
        action = self.view.menu_button.menu().addAction(f"UASG {uasg}", lambda: self.update_table(uasg))


    def fetch_and_create_table(self):
        """Busca os dados da UASG e atualiza o banco de dados."""
        uasg = self.view.uasg_input.text().strip()

        # Verificação se a UASG está vazia ou contém caracteres não numéricos
        if not uasg or not uasg.isdigit():
            QMessageBox.warning(self.view, "Entrada Inválida", "Por favor, insira um número UASG válido.")
            return

        try:
            # Se a UASG já estiver carregada, atualizar os dados
            # Vamos mudar a lógica aqui: se a UASG já existe, apenas carregamos da memória/DB
            # A atualização explícita pode ser uma outra função/botão.
            if uasg in self.loaded_uasgs and self.loaded_uasgs[uasg]: # Verifica se há dados carregados
                self.view.label.setText(f"UASG {uasg} já carregada. Exibindo dados locais.")
                self.view.uasg_input.clear()
                # Não chama update_uasg_data aqui, apenas carrega o que já tem.
                # A função update_table(uasg) já vai pegar os dados de self.loaded_uasgs
                # que foram carregados do DB no __init__ ou após um save.
            else:
                # Se for nova, buscar e salvar
                data = self.model.fetch_uasg_data(uasg)
                if data is None:
                    QMessageBox.critical(self.view, "Erro de API", f"Erro ao buscar dados da UASG {uasg}.")
                    return

                self.model.save_uasg_data(uasg, data)
                self.loaded_uasgs[uasg] = data
                self.add_uasg_to_menu(uasg)
                self.view.label.setText(f"UASG {uasg} carregada e salva com sucesso!")
                # time.sleep(1) # Considere remover ou usar QStatusBar
                self.view.uasg_input.clear()

            # Garante que os dados mais recentes (do DB ou da API) estejam em self.loaded_uasgs
            # antes de chamar update_table
            self.loaded_uasgs = self.model.load_saved_uasgs()

            self.update_table(uasg)
            self.view.tabs.setCurrentWidget(self.view.table_tab)
        except requests.exceptions.RequestException as e: # Exceção mais específica
            QMessageBox.critical(self.view, "Erro de Rede", f"Erro de rede ao buscar UASG {uasg}: {str(e)}")
        except Exception as e: # Genérico para outros erros
            QMessageBox.critical(self.view, "Erro Inesperado", f"Erro inesperado ao processar UASG {uasg}: {str(e)}")

    def delete_uasg_data(self):
        """Deleta os dados da UASG informada e limpa a tabela se ela estiver em uso."""
        uasg_para_deletar = self.view.uasg_input.text().strip()
        if not uasg_para_deletar:
            QMessageBox.warning(self.view, "Entrada Inválida", "Por favor, insira um número UASG para deletar.")
            return

        if uasg_para_deletar not in self.loaded_uasgs:
            QMessageBox.warning(self.view, "Erro", f"Nenhum dado encontrado para a UASG {uasg_para_deletar}.")
            return

        # --- LÓGICA DE VERIFICAÇÃO ADICIONADA AQUI ---
        # Pega o texto do label de informação, ex: "UASG: 787010 - CEIMBRA"
        info_label_text = self.view.uasg_info_label.text()
        
        # Extrai o código da UASG que está sendo exibida na tabela
        uasg_sendo_exibida = ""
        if "UASG: " in info_label_text:
            # Pega a parte depois de "UASG: " e antes do primeiro espaço
            partes = info_label_text.split(" ")
            if len(partes) > 1:
                uasg_sendo_exibida = partes[1]

        # Compara a UASG a ser deletada com a que está na tela
        if uasg_para_deletar == uasg_sendo_exibida:
            self.view.search_bar.clear()
            model = self.view.table.model()
            model.removeRows(0, model.rowCount())
            self.view.uasg_info_label.setText("UASG: -")
            print(f"Limpando a tabela, pois a UASG {uasg_para_deletar} está sendo visualizada.")
            
        
        # Procede com a deleção dos dados do banco de dados
        self.model.delete_uasg_data(uasg_para_deletar)
        self.loaded_uasgs.pop(uasg_para_deletar, None)
        
        # Atualiza o menu de UASGs e limpa o campo de input
        self.load_saved_uasgs()
        self.view.uasg_input.clear()
        
        QMessageBox.information(self.view, "Sucesso", f"UASG {uasg_para_deletar} deletada com sucesso.")

    def update_table(self, uasg):
        """Atualiza a tabela com os dados da UASG selecionada."""
        if uasg in self.loaded_uasgs:
            # Recarrega os dados do arquivo JSON para garantir que estão atualizados
            self.loaded_uasgs = self.model.load_saved_uasgs()
            
            # Atualiza os dados atuais com os dados recarregados
            if uasg in self.loaded_uasgs:
                self.current_data = self.loaded_uasgs[uasg]
                
                # Obter o nome resumido da UASG para mostrar no label
                nome_resumido = ""
                if self.current_data and len(self.current_data) > 0:
                    contrato = self.current_data[0]
                    nome_resumido = contrato.get("contratante", {}).get("orgao", {}).get("unidade_gestora", {}).get("nome_resumido", "")
                
                # Atualiza o label na interface
                self.view.uasg_info_label.setText(f"UASG: {uasg} - {nome_resumido}")
                
                # Popula a tabela com os dados usando a função do módulo controller_table
                populate_table(self, self.current_data)
                self.dashboard_controller.update_dashboard(self.current_data)
                print(f"✅ Tabela atualizada com os dados da UASG {uasg}.")
            else:
                # Limpa o label se não houver dados
                self.view.uasg_info_label.setText(f"UASG: -")
                print(f"⚠ UASG {uasg} não encontrada nos dados recarregados(especifico).")
        else:
            # Limpa o label se a UASG não for encontrada
            self.view.uasg_info_label.setText(f"UASG: -")
            print(f"⚠ UASG {uasg} não encontrada nos dados carregados(geral).")
        self.load_saved_uasgs()

    def clear_table(self):
        """Limpa o conteúdo da tabela."""
        self.view.search_bar.clear()
        model = self.view.table.model()
        model.removeRows(0, model.rowCount())
        
        # Limpa o rótulo da UASG
        self.view.uasg_info_label.setText("UASG: -")
        
        # Limpa o dashboard
        self.dashboard_controller.clear_dashboard()
        
        QMessageBox.information(self.view, "Limpeza", "A tabela foi limpa com sucesso!")

    def show_context_menu(self, position):
        """Exibe o menu de contexto ao clicar com o botão direito na tabela."""
        index = self.view.table.indexAt(position)
        if not index.isValid():
            return

        # Mapeia o índice filtrado para o índice do modelo base
        source_index = self.view.table.model().mapToSource(index)
        row = source_index.row()

        # Certifica-se de que a lista usada está correta
        data_source = self.current_data

        # Verifica se o índice é válido para evitar erro de "index out of range"
        if 0 <= row < len(data_source):
            contrato = data_source[row]
            menu = QMenu(self.view)
            # Adiciona o ícone "init" à ação "Ver Detalhes"
            details_action = menu.addAction(icon_manager.get_icon("init"), "Ver Detalhes")
            details_action.triggered.connect(lambda: self.show_details_dialog(contrato))
            menu.exec(self.view.table.mapToGlobal(position))

    def show_details_dialog(self, contrato):
        """Exibe o diálogo de detalhes do contrato."""
        details_dialog = DetailsDialog(contrato, self.model, self.view) # Passa self.model
        
        # Conectar o sinal data_saved ao método que atualiza a tabela
        details_dialog.data_saved.connect(self.update_table_from_details)
        
        details_dialog.exec()

    def update_table_from_details(self, details_info):
        """Atualiza a tabela quando os dados são salvos na DetailsDialog."""
        update_row_from_details(self, details_info)
    
    def show_msg_dialog(self):
        """
        Abre a janela de geração de mensagens para o contrato selecionado.
        """
        # Pega o índice da linha selecionada na tabela
        selected_indexes = self.view.table.selectionModel().selectedIndexes()
        
        if not selected_indexes:
            QMessageBox.warning(self.view, "Nenhum Contrato Selecionado", "Por favor, selecione um contrato na tabela antes de clicar em Mensagens.")
            return

        # Pega o índice real da fonte de dados
        source_index = self.view.table.model().mapToSource(selected_indexes[0])
        selected_row = source_index.row()
        
        # Pega os dados do contrato selecionado
        contract_data = self.current_data[selected_row]
        
        # Cria e exibe a nova janela de mensagens
        mensagem_controller = MensagemController(contract_data, parent=self.view)
        mensagem_controller.show()

    # Método para abrir a janela de configurações
    def show_settings_dialog(self):
        """Abre a janela de configurações e conecta o sinal de mudança de modo."""
        settings_controller = SettingsController(self.model, self.view)
        settings_controller.mode_changed.connect(self.view.update_status_icon)
        settings_controller.mode_changed.connect(self.view.update_clear_button_icon)
        settings_controller.show()

    def export_status_data(self):
        """Exporta todos os dados de status para um arquivo JSON."""
        all_status_data = self.model.get_all_status_data()
        if not all_status_data:
            QMessageBox.information(self.view, "Exportar Status", "Não há dados de status para exportar.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self.view,
            "Salvar Dados de Status",
            "", # Diretório inicial
            "JSON Files (*.json);;All Files (*)"
        )

        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(all_status_data, f, ensure_ascii=False, indent=4)
                QMessageBox.information(self.view, "Exportar Status", f"Dados de status exportados com sucesso para:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self.view, "Erro ao Exportar", f"Não foi possível salvar o arquivo: {e}")

    def import_status_data(self):
        """Importa dados de status de um arquivo JSON."""
        file_path, _ = QFileDialog.getOpenFileName(
            self.view,
            "Abrir Dados de Status",
            "", # Diretório inicial
            "JSON Files (*.json);;All Files (*)"
        )

        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data_to_import = json.load(f)
                
                self.model.import_statuses(data_to_import)
                QMessageBox.information(self.view, "Importar Status", "Dados de status importados com sucesso!\nA tabela será atualizada.")
                self.load_saved_uasgs() # Recarrega UASGs e atualiza o menu
                # Força a atualização da tabela visível, se houver alguma UASG carregada
                current_uasg_text = self.view.uasg_info_label.text()
                if "UASG: " in current_uasg_text and current_uasg_text.split(" ")[1] != "-":
                    uasg_code = current_uasg_text.split(" ")[1]
                    self.update_table(uasg_code) # Atualiza a tabela com a UASG atual

            except json.JSONDecodeError:
                QMessageBox.critical(self.view, "Erro de Importação", "Arquivo JSON inválido ou corrompido.")
            except Exception as e:
                QMessageBox.critical(self.view, "Erro ao Importar", f"Não foi possível importar os dados: {e}")

    def export_table_to_excel(self):
        """
        Exporta os dados para um arquivo Excel (.xlsx) com formatação avançada,
        incluindo fórmulas dinâmicas e quebra de linha nos cabeçalhos.
        """
        if not self.current_data:
            QMessageBox.information(self.view, "Exportar Tabela", "Não há dados na tabela para exportar.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self.view,
            "Salvar Planilha como...",
            "Relatorio_Contratos.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            # --- PREPARAÇÃO DA PLANILHA ---
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Acordos Administrativos"
            ano_atual = datetime.now().year
            data_atual_str = datetime.now().strftime("%d/%m/%Y")

            # --- LÓGICA PARA INSERIR OS ÍCONES ---
            try:
                # Carrega o ícone da esquerda ("icone.png")
                path_icone = os.path.join('utils', 'icons', 'icone.ico')
                if os.path.exists(path_icone):
                    logo_esquerdo = Image(path_icone)
                    # Ajusta o tamanho do logo para caber nas 3 linhas do cabeçalho
                    logo_esquerdo.height = 70
                    logo_esquerdo.width = 70
                    # Adiciona a imagem na célula A1 (canto superior esquerdo)
                    ws.add_image(logo_esquerdo, 'A1')
                else:
                    print(f"Aviso: Ícone não encontrado em {path_icone}")

                # Carrega o ícone da direita ("acanto.png")
                path_acanto = os.path.join('utils', 'icons', 'acanto.png')
                if os.path.exists(path_acanto):
                    logo_direito = Image(path_acanto)
                    logo_direito.height = 70
                    logo_direito.width = 70
                    # Adiciona a imagem na célula L1 (canto superior direito)
                    ws.add_image(logo_direito, 'L1')
                else:
                    print(f"Aviso: Ícone não encontrado em {path_acanto}")
            
            except Exception as img_error:
                print(f"Ocorreu um erro ao carregar os logos: {img_error}")
                QMessageBox.warning(self.view, "Aviso", "Não foi possível carregar os logos na planilha.")

            # --- CABEÇALHO PRINCIPAL E TÍTULOS ---
            # Mescla as células CENTRAIS para o texto, deixando as colunas A e L para os logos
            ws.merge_cells('B1:K3')
            ws['B1'].value = "CENTRO DE INTENDÊNCIA DA MARINHA EM BRASÍLIA\nDIVISÃO DE OBTENÇÃO"
            ws['B1'].font = Font(bold=True, size=14)
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.merge_cells('A4:L4')
            ws['A4'].value = f"ACORDOS ADMINISTRATIVOS EM VIGOR {ano_atual}"
            ws['A4'].font = Font(bold=True, size=12)
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

            # O restante do seu código permanece exatamente o mesmo
            # --- DATA DE REFERÊNCIA DINÂMICA ---
            ref_date_cell = 'L6'
            ws[ref_date_cell] = f"Data: {data_atual_str}"
            ws[ref_date_cell].font = Font(bold=True, italic=True)
            ws[ref_date_cell].alignment = Alignment(horizontal='center')

            # --- CABEÇALHOS DAS COLUNAS (com quebra de linha) ---
            headers = [
                "SETOR", "MODALIDADE", "N°/ANO", "EMPRESA", "CONTRATO\n- ATA PARECER", 
                "OBJETO", "CELEBRAÇÃO", "TERMO\nADITIVO", "PORTARIA DE\nFISCALIZAÇÃO", 
                "TÉRMINO", "DIAS P/\nVENCIMENTO", "OBS"
            ]
            ws.append(headers)
            
            header_row = ws[7]
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for cell in header_row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # --- DADOS E FÓRMULAS ---
            center_alignment = Alignment(horizontal='center', vertical='center')
            green_font = Font(color="00B050") # Tom de verde padrão do Excel

            for row_idx, contrato in enumerate(self.current_data, start=8): # Começa na linha 8
                data_termino_str = contrato.get("vigencia_fim", "")
                data_termino_excel = None
                if data_termino_str:
                    try:
                        data_termino_excel = datetime.strptime(data_termino_str, "%Y-%m-%d")
                    except ValueError:
                        data_termino_excel = "Data Inválida"

                row_data = [
                    contrato.get("contratante", {}).get("orgao", {}).get("unidade_gestora", {}).get("nome_resumido", "N/A"),
                    contrato.get("modalidade", "N/A"),
                    contrato.get("licitacao_numero", "N/A"),
                    contrato.get("fornecedor", {}).get("nome", ""),
                    contrato.get("numero", ""),
                    contrato.get("objeto", ""),
                    contrato.get("data_assinatura", "N/A"),
                    "XXX",
                    "XXX",
                    data_termino_excel,
                    f'=IF(ISBLANK(J{row_idx}), "N/A", J{row_idx}-TODAY())', # FÓRMULA DINÂMICA
                    ""
                ]
                ws.append(row_data)

                for col_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_alignment
                    
                    if col_idx in [7, 10] and isinstance(cell.value, datetime):
                        cell.number_format = 'DD/MM/YYYY'
                    
                    if col_idx == 11:
                        cell.font = green_font
                        cell.number_format = '0'

            # --- AJUSTE FINAL DAS LARGURAS DAS COLUNAS ---
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 45
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 15
            ws.column_dimensions['L'].width = 20

            workbook.save(file_path)
            QMessageBox.information(self.view, "Exportação Concluída", f"Planilha salva com sucesso em:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self.view, "Erro ao Exportar", f"Ocorreu um erro ao gerar a planilha:\n{e}")  # Funções auxiliares que você precisaria (ou adaptar das existentes em controller_table.py)
    def _calculate_dias_restantes(self, vigencia_fim_str):
        from datetime import date, datetime # Mover imports para o topo do arquivo
        if vigencia_fim_str:
            try:
                vigencia_fim = datetime.strptime(vigencia_fim_str, "%Y-%m-%d").date()
                return (vigencia_fim - date.today()).days
            except ValueError:
                return "Erro Data"
        return "Sem Data"

    def _get_status_for_contrato(self, contrato_id):
        # Lógica similar à de controller_table.py para buscar o status do DB
        if contrato_id and self.model:
            try:
                conn = self.model._get_db_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT status FROM status_contratos WHERE contrato_id = ?", (contrato_id,))
                status_row = cursor.fetchone()
                conn.close()
                if status_row and status_row['status']:
                    return status_row['status']
            except sqlite3.Error:
                return "Erro DB"
        return "SEÇÃO CONTRATOS"

    '''def set_pdf_download_folder(self):
        """Permite ao usuário definir a pasta de download para PDFs."""
        current_path = self.model.load_setting("pdf_download_path", os.path.join(os.path.expanduser("~"), "Downloads"))
        
        folder_path = QFileDialog.getExistingDirectory(
            self.view,
            "Selecionar Pasta para Salvar PDFs",
            current_path
        )
        if folder_path:
            self.model.save_setting("pdf_download_path", folder_path)
            QMessageBox.information(self.view, "Pasta Definida", f"Os PDFs serão salvos em:\n{folder_path}")'''
    
    def populate_previsualization_table(self):
        """Busca os contratos com status diferente de 'SEÇÃO CONTRATOS' e popula a tabela de pré-visualização."""
        data = self.model.get_contracts_with_status_not_default()
        self.view.populate_preview_table(data)
