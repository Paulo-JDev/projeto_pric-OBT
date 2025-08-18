# controller/empenhos_controller.py

from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

class EmpenhoController:
    def __init__(self, model, parent_view=None):
        self.model = model
        self.parent_view = parent_view

    @staticmethod
    def _to_float(value_str):
        if not isinstance(value_str, str):
            return 0.0
        return float(value_str.replace('.', '').replace(',', '.'))

    def generate_report_to_excel(self, contract_data):
        """
        Orquestra a busca de dados, o processamento e a criação do relatório Excel
        com uma aba de resumo e abas detalhadas para cada período.
        """
        contrato_id = contract_data.get("id")

        # 1. Busca os dados necessários
        empenhos, error_empenhos = self.model.get_sub_data_for_contract(contrato_id, "empenhos")
        historico, error_historico = self.model.get_sub_data_for_contract(contrato_id, "historico")

        if error_empenhos or error_historico or not historico:
            QMessageBox.critical(self.parent_view, "Erro de Dados", "Não foi possível buscar o histórico ou os empenhos para gerar o relatório.")
            return

        # 2. Processa os dados para agrupar empenhos por período
        report_data = self._process_data_for_excel(historico, empenhos)

        # 3. Pede ao usuário onde salvar o arquivo
        file_path, _ = QFileDialog.getSaveFileName(
            self.parent_view,
            "Salvar Relatório de Empenhos",
            f"Relatorio_Empenhos_Contrato_{contract_data.get('numero', '').replace('/', '-')}.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        # 4. Cria o arquivo Excel
        try:
            self._create_excel_file(file_path, report_data)
            QMessageBox.information(self.parent_view, "Sucesso", f"Relatório salvo com sucesso em:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self.parent_view, "Erro ao Gerar Excel", f"Ocorreu um erro ao criar a planilha:\n{str(e)}")

    def _process_data_for_excel(self, historico, empenhos):
        """Prepara os dados para o relatório, incluindo a lógica de verificação."""
        periodos = []
        for item in sorted(historico, key=lambda x: x['vigencia_inicio']):
            if item.get("codigo_tipo") in ["50", "55"]:
                periodos.append({
                    "titulo": f"{item['tipo']} {item['numero']}",
                    "inicio": datetime.strptime(item['vigencia_inicio'], "%Y-%m-%d").date(),
                    "fim": datetime.strptime(item['vigencia_fim'], "%Y-%m-%d").date(),
                    "valor_global": self._to_float(item.get('valor_global', '0,00')), # <<< CORRIGIDO
                    "empenhos_do_periodo": []
                })

        for empenho in empenhos:
            try:
                data_emissao = datetime.strptime(empenho['data_emissao'], "%Y-%m-%d").date()
                for periodo in periodos:
                    if periodo['inicio'] <= data_emissao <= periodo['fim']:
                        periodo['empenhos_do_periodo'].append(empenho)
                        break
            except (ValueError, TypeError):
                continue

        report = []
        for periodo in periodos:
            total_empenhado = sum(self._to_float(e.get('empenhado', '0,00')) for e in periodo['empenhos_do_periodo']) # <<< CORRIGIDO
            total_pago = sum(self._to_float(e.get('pago', '0,00')) for e in periodo['empenhos_do_periodo']) # <<< CORRIGIDO

            obs = "OK"
            if total_empenhado > periodo['valor_global'] or total_pago > periodo['valor_global']:
                obs = "Sera??"

            report.append({
                "titulo": periodo['titulo'],
                "inicio": periodo['inicio'].strftime("%d/%m/%Y"),
                "fim": periodo['fim'].strftime("%d/%m/%Y"),
                "valor_global": periodo['valor_global'],
                "total_empenhado": total_empenhado,
                "total_pago": total_pago,
                "obs": obs,
                "empenhos_detalhados": periodo['empenhos_do_periodo']
            })
        
        return report

    def _create_excel_file(self, file_path, report_data):
        """Cria e formata o arquivo .xlsx com a aba de resumo e as abas detalhadas."""
        workbook = Workbook()
        
        # --- ESTILOS GERAIS ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Estilos para o corpo da tabela (dark mode)
        body_font = Font(color="FFFFFF")
        body_fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
        
        # Estilos para a seção BASE
        base_header_font = Font(bold=True, color="FFFFFF")
        base_header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        base_label_font = Font(bold=True, color="FFFFFF") # Letra branca para os labels da base

        separator_border = Border(bottom=Side(style='thick', color="4F81BD"))

        # --- 1. ABA DE RESUMO GERAL ---
        ws_resumo = workbook.active
        ws_resumo.title = "Resumo Geral"
        
        headers_resumo = ["Período", "Vigência", "Valor Global", "Total Empenhado", "Total Pago", "OBS"]
        ws_resumo.append(headers_resumo)
        for cell in ws_resumo[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row_data in report_data:
            ws_resumo.append([
                row_data['titulo'], f"{row_data['inicio']} a {row_data['fim']}",
                row_data['valor_global'], row_data['total_empenhado'],
                row_data['total_pago'], row_data['obs']
            ])

        for row in ws_resumo.iter_rows(min_row=2, max_row=ws_resumo.max_row):
            for cell in row:
                cell.font = body_font
                cell.fill = body_fill
            
            row[0].alignment, row[1].alignment, row[5].alignment = center_align, center_align, center_align
            row[2].number_format = '"R$" #,##0.00'
            row[3].number_format = '"R$" #,##0.00'
            row[4].number_format = '"R$" #,##0.00'
            if row[5].value == "Sera??":
                row[5].font = Font(bold=True, color="FF0000")

        for i, col in enumerate(ws_resumo.columns):
            ws_resumo.column_dimensions[get_column_letter(i + 1)].auto_size = True

        # --- 2. ABAS DETALHADAS PARA CADA PERÍODO ---
        for periodo_data in report_data:
            sheet_title = re.sub(r'[\\/*?:\[\]]', '', periodo_data['titulo'])[:31]
            ws_detalhe = workbook.create_sheet(title=sheet_title)

            # --- SEÇÃO "BASE" ---
            ws_detalhe.merge_cells('A1:C1')
            base_title_cell = ws_detalhe['A1']
            base_title_cell.value = "INFORMAÇÕES DE BASE DO PERÍODO"
            base_title_cell.font = base_header_font
            base_title_cell.fill = base_header_fill
            base_title_cell.alignment = Alignment(horizontal='center')

            for row_num in range(2, 5):
                for col_num in range(1, 4):
                    cell = ws_detalhe.cell(row=row_num, column=col_num)
                    cell.font = body_font
                    cell.fill = body_fill
            
            ws_detalhe['A2'] = "Período de Vigência:"
            ws_detalhe['A2'].font = base_label_font
            ws_detalhe['B2'] = f"{periodo_data['inicio']} a {periodo_data['fim']}"
            
            ws_detalhe['A3'] = "Valor Global do Período:"
            ws_detalhe['A3'].font = base_label_font
            ws_detalhe['B3'] = periodo_data['valor_global']
            ws_detalhe['B3'].number_format = '"R$" #,##0.00'
            
            ws_detalhe['A4'] = "Natureza da Despesa:"
            ws_detalhe['A4'].font = base_label_font
            natureza_despesa = "N/A"
            if periodo_data['empenhos_detalhados']:
                natureza_despesa = periodo_data['empenhos_detalhados'][0].get('naturezadespesa', 'N/A')
            ws_detalhe['B4'] = natureza_despesa
            ws_detalhe.merge_cells('B4:C4')

            for col_num in range(1, 9):
                 ws_detalhe.cell(row=4, column=col_num).border = separator_border
            
            # --- CABEÇALHOS DA TABELA DE EMPENHOS (LINHA 6) ---
            headers_detalhe = [
                "Número Empenho", "Data Emissão", "Valor Empenhado", 
                "Valor a Liquidar", "Valor Pago", "Documento de Pagamento"
            ]
            # Adiciona os cabeçalhos na linha 6 (a próxima linha vazia após a linha 5 de espaço)
            ws_detalhe.append(headers_detalhe)
            
            # --- CORREÇÃO APLICADA AQUI ---
            # O loop agora aplica o estilo na linha correta (linha 6)
            for cell in ws_detalhe[6]:
                cell.font = base_header_font # Fonte branca e negrito
                cell.fill = base_header_fill   # Fundo azul escuro (igual ao título)
                cell.alignment = center_align

            # Preenche com os empenhos detalhados
            for empenho in periodo_data['empenhos_detalhados']:
                doc_pagamento_url = empenho.get("links", {}).get("documento_pagamento")
                numero_empenho = empenho.get('numero', 'N/A')
                
                row_to_append = [
                    numero_empenho,
                    datetime.strptime(empenho.get('data_emissao'), "%Y-%m-%d").date() if empenho.get('data_emissao') else "N/A",
                    self._to_float(empenho.get('empenhado', '0,00')),
                    self._to_float(empenho.get('aliquidar', '0,00')),
                    self._to_float(empenho.get('pago', '0,00')),
                    f'=HYPERLINK("{doc_pagamento_url}", "{numero_empenho}_linkdoc")' if doc_pagamento_url else "Sem link"
                ]
                ws_detalhe.append(row_to_append)

            # Formata as células da tabela (começando da linha 7)
            for row in ws_detalhe.iter_rows(min_row=7, max_row=ws_detalhe.max_row):
                for cell in row:
                    cell.font = body_font
                    cell.fill = body_fill
                
                row[0].alignment = center_align
                row[1].alignment = center_align
                row[1].number_format = 'DD/MM/YYYY'
                row[2].number_format = '"R$" #,##0.00'
                row[3].number_format = '"R$" #,##0.00'
                row[4].number_format = '"R$" #,##0.00'
                
                if row[5].value != "Sem link":
                    row[5].font = Font(color="5999FF", underline="single")
                    row[5].alignment = center_align

            for i, col in enumerate(ws_detalhe.columns):
                ws_detalhe.column_dimensions[get_column_letter(i + 1)].auto_size = True
        
        workbook.save(file_path)
